import warnings
from cryptography.utils import CryptographyDeprecationWarning
from commons import __get_parameters
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore", category=CryptographyDeprecationWarning)
import logging
from commons import *
from bot_base import *
import pandas as pd
from retrying import retry
from wrapt_timeout_decorator import *
from datetime import timedelta
from playwright.sync_api import sync_playwright, expect
from bs4 import BeautifulSoup
import urllib.parse
from lxml import etree
import pandas as pd
from openpyxl import load_workbook
from http import HTTPStatus
from tinydb import TinyDB, Query
import humanize
import base64
import ssl
import certifi
import traceback
from collections import Counter

colunas = {valor: indice for indice, valor in enumerate(pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="ids", header=None).iloc[:, 0].tolist())}

@timeit
@handle_exceptions(default_return=[])
@retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
def get_protocols(session) -> list:

    payload = json.dumps({
        "jsonrpc": "2.0",
        "id": 1,
        "method": "timesync"
    })
    
    response = safe_post(global_parameters["mobilesaude.timestamp"], headers={'Content-Type': 'application/json'}, data=payload, verify=False)

    json_object = json.loads(response.text)

    payload = {
        "draw": 1,
        "columns[0][data]": 0,
        "columns[0][name]": None,
        "columns[0][searchable]": True,
        "columns[0][orderable]": False,
        "columns[0][search][value]": None,
        "columns[0][search][regex]": False,
        "order[0][column]": 0,
        "order[0][dir]": "asc",                     
        "start": 0,
        "length": 9999, # global_parameters["mobilesaude.total_records"],
        "search[value]": None,
        "search[regex]": False,
        "usuario": 0,
        "atribuicao": global_parameters["mobilesaude.atribuicao"],
        "prestador_plano": None,
        "status": global_parameters[f'mobilesaude.status_filtro_{global_parameters["env"]}'],
        "prestador_estado": None,
        "prestador_cidade": None,
        "data_solicitacao_inicio": None,
        "data_solicitacao_fim": None,
        "matricula": None,
        "_": json_object["result"]      
    }                 

    query_string = urllib.parse.urlencode({k: "" if v is None else v for k, v in payload.items()})

    custom_headers = session.headers.copy()
    custom_headers["x-requested-with"] = "XMLHttpRequest"
    custom_headers["Accept"] = "application/json, text/javascript, */*; q=0.01"
    custom_headers["User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36"

    logging.info(f'Pesquisando protocolos...')

    # 
    protected_page_url = global_parameters["mobilesaude.query_protocolos"] %(query_string)
    protected_response = session.get(protected_page_url, headers=custom_headers, verify=False)

    # check if the request was successful
    if not response.status_code == HTTPStatus.OK:
        raise Exception(f"Falha ao recuperar a lista de protocolos: {response.status_code}")
    
    protocols = json.loads(protected_response.text)["data"]

    sorted_protocols = sorted(
        protocols,
        key=lambda x: x[1]
    )

    with open(f'{log}/{today}{today_.strftime("%H%M%S")}.json', "w") as file:
        json.dump(sorted_protocols, file, indent=4)

    return sorted_protocols

@timeit
@handle_exceptions(default_return=False)
@retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)    
def login_mobile(session) -> bool:

    payload = {
        "email": config["mobilesaude"]["username"],
        "senha": config["mobilesaude"]["password"]
    }

    response = session.post(global_parameters["mobilesaude.submit_login"], data=payload, verify=False)

    # check if the request was successful
    if not response.status_code == HTTPStatus.OK:
        logging.error(f"Erro no login: {response.status_code}")   
        return False 

    logging.info(f'Login com sucesso {config["mobilesaude"]["username"]}...')

    # SAUDE AMS
    logging.info(f'Navegando para SAUDE AMS...')

    response = session.get(global_parameters[f'mobilesaude.saude_ams_{global_parameters["env"]}'], verify=False)            

    # check if the request was successful
    if not response.status_code == HTTPStatus.OK:
        logging.error(f"Falha na navegação {response.status_code}")   
        return  False   
    
    return True
    
@timeit
def export_backlog():

    # Create a session object to maintain cookies and headers across requests
    session = requests.Session()

    if not login_mobile(session):
        return None
    
    protocols = get_protocols(session)

    # Extração das datas (coluna índice 4)
    datas = [linha[4] for linha in protocols]

    # Contar ocorrências por data
    contagem = Counter(datas)
    total = sum(contagem.values())

    # Exibir resultados
    registros = []
    logging.info(f"{'Data':<15} {'Qtd':<5} {'%':<6}")
    for data, qtd in sorted(contagem.items()):
        percentual = round((qtd / total) * 100, 2)
        registros.append({
            "Data": data,
            "Quantidade": qtd,
            "Percentual (%)": percentual
        })
        logging.info(f"{data:<15} {qtd:<5} {percentual:>5.1f}%")

    # Exportar para Excel
    df = pd.DataFrame(registros)
    df.to_excel(f'{log}/backlog_por_data_{today}{today_.strftime("%H%M%S")}.xlsx', index=False) 

    return None    

@timeit
def read_protocols_mobile_saude() -> str:
        
    @timeit
    def process_protocol() -> list:

        file_id = os.path.basename(file)
        env = global_parameters["env"]
        refund_id = match(r'(?<=value=")(.*?)(?=")', protocols[protocols_count][0]).strip()
        protocol_id = protocols[protocols_count][1]
        status_id = match(r'(?<=data-status=")(.*?)(?=")', protocols[protocols_count][0]).strip()
        refund_type = protocols[protocols_count][2]
        refund_qty = protocols[protocols_count][3]
        protocol_date = protocols[protocols_count][4]
        refund_value = protocols[protocols_count][5]
        status_desc = re.sub(r"<[^>]*>", "", protocols[protocols_count][6])    
        PEG = ""
        holder_name = ""
        phone_number = ""
        holder_cpf = ""
        plan = ""
        payment_day = ""
        payment_type = ""
        lot = ""
        notes = ""
        expense_id = ""
        card = ""
        user = ""
        expense_status = ""
        supplier_id = ""
        supplier_name = ""
        supplier_state = ""
        supplier_city = ""
        expense_date = ""
        expense_nf = ""
        guide_number = ""        
        assigned = ""
        comment = ""
        complement = ""

        logging.info(f'Extraindo dados do protocolo {protocol_id}...')

        #
        response = session.get( global_parameters["mobilesaude.reembolso"] %(refund_id), verify=False)            

        # check if the request was successful
        if response.status_code == HTTPStatus.OK:

            #
            soup = BeautifulSoup(response.text, "html.parser")
            dom = etree.HTML(str(soup))

            #
            PEG = match(r'\d{2}\.\d{3}\.\d{3}', safe_soup_find(soup,"textarea", {"name":"observacao_interna"}, 'text') or "")   
            PEG = safe_soup_find(soup,"input", {"name":"numero_lote", "type": "text"},'value') or PEG
            holder_name = safe_soup_find(soup,"input", {"name":"nome_titular", "type": "hidden"},'value')
            phone_number = safe_soup_find(soup,"input", {"name":"telefone", "type": "hidden"},'value')
            holder_cpf = safe_soup_find(soup,"input", {"name":"cpf_titular", "type": "hidden"},'value')
            plan = safe_xpath(dom, '//label[@for="tipo-reembolso"]/following-sibling::p/text()')
            payment_day = safe_soup_find(soup,"input", {"name":"data_pagamento", "type": "text"},'value')   
            payment_type = safe_xpath(dom,'//select[@name="despesa[1][desembolso]"]/option[@selected]/text()')
            lot = safe_soup_find(soup,"input", {"name":"numero_lote", "type": "text"},'value')   
            notes = safe_soup_find(soup,"textarea", {"name":"observacao"}, 'text')
            expense_id = safe_soup_find(soup,"input", {"name":"despesa[1][id_despesa_reembolso]", "type": "hidden"},'value')
            card = safe_soup_find(soup,"input", {"name":"despesa[1][utilizador_matricula]", "type": "hidden"},'value')
            user = (safe_xpath(dom, '//label[normalize-space(.)="CARTÃO DO UTILIZADOR"]/following-sibling::p[1]/text()')).split("-", 1)[1].strip()
            expense_status = safe_xpath(dom,'//select[@name="despesa[1][id_status_reembolso]"]/option[@selected]/text()')
            supplier_id = safe_xpath(dom,'//label[@for="documento"]/following-sibling::p/text()')
            supplier_name = safe_xpath(dom,'//label[@for="nome_fantasia"]/following-sibling::p/text()')
            supplier_state = safe_xpath(dom,'//label[text()="ESTADO DO PRESTADOR"]/following-sibling::p/text()')
            supplier_city = safe_xpath(dom,'//label[text()="CIDADE DO PRESTADOR"]/following-sibling::p/text()')
            expense_date = safe_xpath(dom,'//label[text()="DATA DA DESPESA"]/following-sibling::p/text()')
            expense_nf = safe_xpath(dom,'//label[text()="NÚMERO DA NOTA FISCAL / RECIBO"]/following-sibling::p/text()')
            assigned = safe_xpath(dom, '//p[text()="Atribuição: "]/strong/text()')
            comment = "PEG previamente associado" if PEG else ""

        else: 
            comment = f"Erro ao abrir detalhes do protocolo {protocol_id} code: {response.status_code}"
        
        if comment:
            logging.error(comment)

        return [
            file_id,
            env,
            refund_id,
            protocol_id,
            protocol_date,
            status_id,
            status_desc,
            refund_type,
            refund_qty,
            refund_value,
            card,
            user,
            holder_name,
            holder_cpf,
            phone_number,
            plan,
            payment_day,
            payment_type,
            lot,
            expense_id,
            expense_status,
            supplier_id,
            supplier_name,
            supplier_state,
            supplier_city,
            expense_date,
            expense_nf,
            guide_number,
            PEG,
            notes,
            assigned,
            comment,
            complement
        ]
    
    @timeit  
    def get_protocols_by_env(env) -> pd.DataFrame:

        logging.info(f'Selecionando protocolos do ambiente {env}...')    

        db = LMDBWrapper()

        # Converter para DataFrame
        df = pd.DataFrame(db.all())

        if not df.empty:

            # Ordenar por datetime decrescente
            df = df.sort_values(by='file_id', ascending=False)    

            # Pegar a primeira ocorrência de cada protocol_id (a mais recente)
            df = df.drop_duplicates(subset='protocol_id', keep='first')    

            # Filtrar apenas registros do ambiente
            df = df[df['env'] == env]

        else:
            
            df = pd.DataFrame(columns=column_ids)
            
        logging.info(f'Selecionando {len(df)} protocolos do ambiente {env}... OK')    

        return df        
    
    @timeit    
    def get_last_protocols_with_error() -> pd.DataFrame:

        def extrair_datetime(file_id):
            try:
                # Procura padrão de data + hora no formato 20250413_1652
                match = re.search(r'(\d{8}_\d{4})', file_id)
                if match:
                    return datetime.strptime(match.group(1), "%Y%m%d_%H%M")
            except:
                pass
            return None   

        logging.info(f'Selecionando protocolos com erro nas últimas {global_parameters["sla_tratamento_erros"]} horas...')    

        df = get_protocols_by_env(global_parameters["env"])  

        # cria a coluna com datetime
        df['file_datetime'] = df['file_id'].apply(extrair_datetime)

        # Limites de tempo
        agora = datetime.now()
        limite_inferior = agora - timedelta(hours=int(global_parameters["sla_tratamento_erros"]))       

        # Filtro
        filtro = (
            df['PEG'].isnull() &
            df['file_datetime'].notnull() &
            (df['file_datetime'] >= limite_inferior) &
            (df['file_datetime'] <= agora) &
            (~(df['comment'] == "OK")) &
            df['comment'].notnull()
        )        

        df = df[filtro]

        # dropa as colunas que não estão em column_names
        df = df[["protocol_id"]]

        # Ordenar do mais antigo para o mais novo
        df = df.sort_values(by='protocol_id', ascending=True)        

        total = len(df)

        for i, row in enumerate(df.itertuples(index=False), start=1):
            logging.info(f'Protocolo {row.protocol_id} com erro {i}/{total}')    

        logging.info(f'{total} protocolo(s) selecionado(s) com erro.')            

        return df
    
    @timeit  
    def get_reprocess_protocols() -> pd.DataFrame:

        logging.info(f'Selecionando protocolos para reprocessamento...')    

        df = get_protocols_by_env(global_parameters["env"])  

        df = df[
            (df['PEG'].notnull()) &
            (~(df['comment'] == "OK"))
        ]

        # Limpa as mensagens de erro para poder reprocessar o protocolo
        df['comment'] = None

        #
        df.rename(columns=dict(zip(column_ids, column_names)), inplace=True)

        # dropa as colunas que não estão em column_names
        df = df[column_names]

        # Ordenar do mais antigo para o mais novo
        df = df.sort_values(by='protocolo', ascending=True)

        total = len(df)

        for i, row in enumerate(df.itertuples(index=False), start=1):
            logging.info(f'Protocolo {row.protocolo} do arquivo {row.arquivo} selecionado para reprocessamento {i}/{total}')    

        logging.info(f'{total} protocolo(s) selecionado(s) para reprocessamento.')    

        return df        
    
    logging.info(f'Extraindo protocolos do Mobilesaude...')
    logging.info(f'Acessando o site...')
    logging.info(f'{global_parameters["mobilesaude.url"]}')        
    logging.info(f'Usuário {config["mobilesaude"]["username"]}')   
     
    # Create a session object to maintain cookies and headers across requests
    session = requests.Session()

    if not login_mobile(session):
        return None
    
    file = f'{log}/protocolos_{today}_{today_.strftime("%H%M")}.xlsx'
    array = [pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="labels", header=None).iloc[:, 0].tolist()]
    column_names = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="labels", header=None).iloc[:, 0].tolist()
    column_ids = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="ids", header=None).iloc[:, 0].tolist()

    df_reprocess_protocols = get_reprocess_protocols()
    protocols = set(sorted(df_reprocess_protocols['protocolo'].values))

    # somente pesquisar protocolos se a lista de reprocessamento for menor que a quantidade esperada para processamento
    if len(df_reprocess_protocols) <= int(global_parameters["mobilesaude.total_records"]):

        # lista de protocolos que apresentaram erro nas últimas x horas
        df_protocols_with_error = get_last_protocols_with_error()
        protocols_with_error = set(sorted(df_protocols_with_error['protocol_id'].values))s
        
        # todos os protocolos pendentes de atuação existentes no mobile
        protocols = get_protocols(session)
        protocols = set(sorted([[item[1] for item in protocols]]))

        # todos os protocolos para reprocessamento
        reprocess_protocols = set(sorted(df_reprocess_protocols['protocolo'].values))

        # Apenas elementos de set2 que não estão em set1
        protocols = protocols_with_error - protocols

        # Apenas elementos de set2 que não estão em set1
        protocols = protocols - reprocess_protocols

        # Unir com set1
        protocols = reprocess_protocols | protocols

        #
        protocols = set(sorted(protocols)[:int(global_parameters["mobilesaude.total_records"])])






        # protocols_count = 0

        # while len(array)-1 < int(global_parameters["mobilesaude.total_records"]) and protocols_count <= len(protocols) -1: 
        
        #     # verifica se o protocolo já está na lista de reprocessamento
        #     if not protocols[protocols_count][1] in df_reprocess_protocols['protocolo'].values:
                
        #         logging.info(f'Adicionando protocolo {protocols[protocols_count][1]} {len(array)-1}/{int(global_parameters["mobilesaude.total_records"])}...')
        #     else:
        #         logging.info(f'Protocolo {protocols[protocols_count][1]} já estava na lista de reprocessamento')  
            
        #     protocols_count += 1

    for protocol in protocols:

        array.append(process_protocol())
    
        logging.info(f'Adicionando protocolo {protocol} {len(array)-1}/{int(global_parameters["mobilesaude.total_records"])}...')
    
    
    logging.info(f'Exportando {file}...')

    # Criar um DataFrame com os dados
    df = pd.DataFrame(array[1:], columns=array[0]).astype(str)

    df_reprocess_protocols['arquivo'] = os.path.basename(file)

    df = pd.concat([df_reprocess_protocols, df], ignore_index=True)

    # Limitar a quantidade de protocolos
    df = df.head(int(global_parameters["mobilesaude.total_records"]))    

    # Ordenar do mais antigo para o mais novo
    df = df.sort_values(by='protocolo', ascending=True)

    # Salvar corretamente em formato aberto para `openpyxl`
    df.to_excel(file, index=False, engine="openpyxl")        

    return file

@timeit
def create_peg_benner(worksheet) -> list:
 
    logging.info(f'Lançando PEGs no Benner...')

    protocols_total = 0
    protocols_count = 0
    protocols_success = 0          

    #
    refund_types = __get_parameters(sheet_name="refund_types", key="from", value="to")

    with sync_playwright() as playwright:

        @timeit
        @handle_exceptions(default_return=False)
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def exists(sc, label, value, dependValueList="") -> bool:  
            return len(search(sc, label, value, dependValueList)) > 0
    
        @timeit
        @handle_exceptions(default_return=[])
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def search(sc, label, value, dependValueList="") -> list:  

            logging.info(f'Pesquisando {value} em {label}...')       

            payload = {
                "query": str(value),
                "sc": sc,
                "dependValueList": dependValueList,
                "maxRows": 1,
                "startRow": 0,
                "fieldsJson": ""
            }

            # Usar o contexto com sessão ativa
            request_context = context.request

            response = request_context.post(
                global_parameters[f'benner.api_search_{global_parameters["env"]}'],
                form=payload,
                ignore_https_errors=True
            )                

            # check if the request was successful
            if response.status != HTTPStatus.OK:
                logging.error(f"{response.text()}")  
                return []

            result = json.loads(response.text())

            if len(result) == 0:
                logging.error(f'{value} não encontrado')  
                return []

            return result
    
        @timeit
        def add_peg() -> bool:     

            @timeit
            def fill_guide() -> str:

                nonlocal guide_number, executor_sc, endereco_executor_sc

                @timeit
                def fill_principal_guide():

                    nonlocal guide_number

                    logging.info(f'Preenchendo guia principal...')
                    
                    # preencher o número da Ordem
                    safe_locator(page,"xpath=//*[@data-label='Ordem']//input[@type='text']").first.fill(row[colunas.get("refund_qty")].value)

                    guide_number = safe_locator(page,"xpath=//*[@data-label='N&#250;mero da guia']//input[@type='text']").first.input_value().strip()
                    
                    # copiar o número da guia para Número da Guia Principal
                    safe_locator(page,"xpath=//*[@data-label='N&#250;mero da Guia Principal']//input[@type='text']").first.fill(guide_number)

                    # copiar o número da guia para Nr Guia Solicitação
                    safe_locator(page,"xpath=//*[@data-label='Nr Guia Solicita&#231;&#227;o']//input[@type='text']").first.fill(guide_number)

                    # copiar o número da guia para No Guia Prestador
                    safe_locator(page,"xpath=//*[@data-label='N&#186; Guia Prestador']//input[@type='text']").first.fill(guide_number)

                @timeit
                def fill_data_guide() -> str:

                    @timeit                 
                    def set_executor_address() -> str:        

                        element = page.locator('select[data-fieldname="ENDERECOEXECUTOR"]') 
                        
                        page.evaluate("""
                            (id) => {
                                Benner.Apps.CustomLookup.showDialog(id);return false;
                            }  
                        """, element.get_attribute("id"))   

                        wait_for_load_state(page, 1000)               
                        
                        # Aguardar carregamento do iframe (opcional, se necessário)
                        page.wait_for_selector("iframe")

                        # Acessar o primeiro iframe da página
                        frames = page.frames
                        frame = frames[1]

                        if not is_element_ready(frame, "#Resultado_SimpleGrid > tbody > tr", timeout=int(global_parameters["timeout"])*2):
                            return f'Nenhum endereço encontrado para o prestador'                        

                        first_row = frame.locator("#Resultado_SimpleGrid > tbody > tr").first
                        first_row.click()

                        return None

                    nonlocal executor_sc, endereco_executor_sc

                    logging.info(f'Criando guia dados da guia...')

                    supplier_id = formatar_cpf_cnpj(row[colunas.get("supplier_id")].value)

                    # clicar na aba Dados da Guia
                    safe_locator(page,'//a[@data-toggle="tab" and normalize-space(text())="Dados da Guia"]').click()

                    wait_for_load_state(page)

                    # recupera o data-searchcontext do EXECUTOR para facilitar consultas futuras
                    if not executor_sc:
                        locator = safe_locator(page, 'select[data-fieldname="EXECUTOR"]')
                        if locator.count():
                            executor_sc = locator.get_attribute('data-searchcontext')       

                    if not endereco_executor_sc:
                        locator = safe_locator(page, 'select[data-fieldname="ENDERECOEXECUTOR"]')
                        if locator.count():
                            endereco_executor_sc = locator.get_attribute('data-searchcontext')                                                     

                    # Geral > Beneficiário
                    if not fill_select2(page, 'select[data-fieldname="BENEFICIARIO"]', row[colunas.get("user")].value):
                        return "Beneficiário não encontrado ou inválido"

                    # Geral > Executor
                    if not fill_select2(page, 'select[data-fieldname="EXECUTOR"]', supplier_id):
                        return "Prestador não encontrado ou inválido"
                    
                    # Geral > Endereço executor
                    if not fill_select2_by_index(page, 'select[data-fieldname="ENDERECOEXECUTOR"]', 1):
                        if not fill_select2(page, 'select[data-fieldname="ENDERECOEXECUTOR"]', ""):
                            if (error_message := set_executor_address()):
                                return error_message                    

                    return None

                @timeit
                def fill_service_characteristics_guide():

                    logging.info(f'Preenchendo guia características do atendimento...')

                    # clicar na aba Dados da Guia
                    safe_locator(page,'//a[@data-toggle="tab" and normalize-space(text())="Características do Atendimento"]').click()

                    wait_for_load_state(page)

                    # Finalidade do atendimento = Rotina
                    fill_select2(page, 'select[data-fieldname="FINALIDADEATENDIMENTO"]', global_parameters["benner.finalidade_atendimento"]) 

                    # Local atendimento = Rede Livre Escolha
                    selector = f'input[value="{global_parameters["benner.local_atendimento"]}"]'
                    if not is_element_ready(page, selector, printscreen=False):
                        fill_select2(page, 'select[data-fieldname="LOCALATENDIMENTO"]', global_parameters["benner.local_atendimento"]) 

                    # selector = f'input[value="{global_parameters["benner.local_atendimento"]}"]'
                    # input_element = safe_locator(page,selector)

                    # if not input_element.count() > 0 and input_element.first.is_visible():
                    #     fill_select2(page, 'select[data-fieldname="LOCALATENDIMENTO"]', global_parameters["benner.local_atendimento"]) 

                    # Regime de atendimento = Ambulatorial
                    fill_select2(page, 'select[data-fieldname="REGIMEATENDIMENTO"]', global_parameters["benner.regime_atendimento"]) 

                    # Condição de atendimento = Eletivo
                    fill_select2(page, 'select[data-fieldname="CONDICAOATENDIMENTO"]', global_parameters["benner.condicao_atendimento"]) 

                    # Tipo do tratamento
                    fill_select2(page, 'select[data-fieldname="TIPOTRATAMENTO"]', global_parameters["benner.tipo_tratamento"]) 

                    # Objetivo do tratamento
                    fill_select2(page, 'select[data-fieldname="OBJETIVOTRATAMENTO"]', global_parameters["benner.objetivo_tratamento"]) 
               
                logging.info(f'Criando guia...')
                
                # cliar em GUIAS
                safe_locator(page,f'xpath=//div/span[normalize-space(text())="Guias"]').first.click()   

                wait_for_load_state(page)   

                # clica em +Novo
                with page.expect_navigation():
                    safe_locator(page,"//div[contains(@class, 'portlet light')][.//span[normalize-space(text())='Guias']]//a[normalize-space(text())='Novo']").click()

                error_message = get_error_message()
                if error_message:
                    return f'Não foi possível criar uma nova guia. ' + re.sub('\s+',' ', error_message)

                steps = [
                    (fill_principal_guide, ()),
                    (fill_data_guide, ()),
                    (fill_service_characteristics_guide, ())
                ]

                for step_func, args in steps:
                    
                    error_message = step_func(*args)
                    
                    if error_message is not None:
                        return error_message 

                # Salvar
                page.keyboard.press("Control+Enter")

                wait_for_load_state(page, 1000)

                error_message = get_error_message()
                if error_message:
                    return f'Não foi possível salvar a guia. ' + re.sub('\s+',' ', error_message)

                logging.info(f"Guia {guide_number} lançada com sucesso")

                return None

            @timeit
            def fill_peg() -> str:

                logging.info(f'Preenchendo PEG...')

                nonlocal peg, client_id
                expense_nf = ''.join(filter(str.isdigit, row[colunas.get("expense_nf")].value or ""))
                expense_date = formatar_data(row[colunas.get("expense_date")].value or "")

                result = navigate_postback_using_form_data(page, form_data, action_url, "ctl00$Main$WDG_V_FILIAIS_779_LNK1", "New")

                if not result["success"]:
                    return f'Falha ao tentar criar o PEG: {result.get("error")}'

                # Selecione "Em Digitação"

                # xpath = "//a[@title='Ctrl + Insert']"

                # if not safe_locator(page, xpath).is_visible():
                #     safe_locator(page,f'//div/span[normalize-space(text())="Em Digitação"]').click()
                #     wait_for_load_state(page)               

                # if not is_element_ready(page, xpath):
                #     return f'Botão "Ctrl + Insert" não está visível'             
                
                # with page.expect_navigation():
                #     safe_locator(page, xpath).first.click()      

                client_id = get_client_id(page)

                if not client_id:
                    return f'client_id não encontrado no corpo da página'   

                # Tipo de PEG = Reembolso
                safe_locator(page,"xpath=//label[normalize-space(text())='Reembolso']").first.click() 

                # Data de recebimento - protocol_date
                safe_locator(page,"xpath=//div[@data-label='Recebimento']//input[@type='text']").first.fill(formatar_data(row[colunas.get("protocol_date")].value))
                
                # Despesa - expense_nf
                if expense_nf:
                    input_locator = safe_locator(page,"xpath=//span[@data-label='N&#250;mero da nota']//input[@type='text']") 
                    input_locator.focus()
                    for letra in expense_nf:
                        input_locator.press(letra)
                    input_locator.press("Tab")

                    # Data de emissão da NF = expense_date
                    safe_locator(page,"xpath=//div[@data-label='Data Emiss&#227;o Nota']//input[@type='text']").first.fill(expense_date)

                # Quantidade de Guias – Apresentada = refund_qty
                page.evaluate('''
                    (value) => {
                        $("span[data-field='QTDGUIA'] input[type='text']").val(value).trigger("change");
                    }
                ''', row[colunas.get("refund_qty")].value)

                # Tipo do PEG
                if not row[colunas.get("refund_type")].value in refund_types:
                    peg_type = global_parameters["benner.default_peg_type"]
                else:
                    peg_type = refund_types[row[colunas.get("refund_type")].value]

                fill_select2(page, 'select[data-fieldname="TIPOPEG"]', peg_type) 

                # Beneficiário Titular 
                fill_select2(page, 'select[data-fieldname="BENEFICIARIO"]', row[colunas.get( global_parameters["benner.chave_beneficiario"] )].value) 

                # Gravando o protocolo na observação
                safe_locator(page,'//a[@data-toggle="tab" and normalize-space(text())="Observações"]').click()
                observacao_interna = safe_locator(page,'div:has(.label-title:has-text("Observação")) textarea').first
                observacao_interna.fill(f'Protocolo mobilesaude:[{row[colunas.get("protocol_id")].value}]\n{observacao_interna.input_value()}')

                # volta para a 1a aba
                safe_locator(page,'//a[@data-toggle="tab" and normalize-space(text())="Principal"]').click()
                
                # Salvar
                page.keyboard.press("Control+Enter")

                wait_for_load_state(page, 1000)

                error_message = get_error_message()
                if error_message:
                    return f'Não foi possível salvar o formulário. ' + re.sub('\s+',' ', error_message)

                peg = get_peg()

                logging.info(f'PEG {peg} lançado com sucesso')
                
                return None

            def valid_required_fields() -> str:

                logging.info(f'Validando campos obrigatórios...')

                supplier_id = formatar_cpf_cnpj(row[colunas.get("supplier_id")].value)
                isCPF = is_CPF(supplier_id)
                isCNPJ = is_CNPJ(supplier_id)
                expense_nf = ''.join(filter(str.isdigit, row[colunas.get("expense_nf")].value or ""))
                expense_date = formatar_data(row[colunas.get("expense_date")].value or "")

                if not row[colunas.get("user")].value:
                    return "Nome do beneficiário não informado"

                # código do prestador deve ser um CPF ou CNPJ válido
                if not isCPF and not isCNPJ:
                    return "Código do prestador inválido"
                
                # Data da NF é obrigatório para prestador PJ
                if isCNPJ and not expense_date:
                    return "Data da NF não informada ou inválida"                  

                # Número da NF é obrigatório para prestador PJ
                if isCNPJ and not expense_nf:
                    return "NF não informada ou inválida"             
                
                # Número da NF deve estar acompanhada da data
                if (expense_nf and not expense_date):
                    return "Número da NF deve estar acompanhada da data"  
                
                # Verifica se o prestador existe na base
                if executor_sc:

                    result = search(executor_sc, "prestador", supplier_id)

                    if not result:
                        return 'Prestador não encontrado ou inválido'
                    else:
                        if endereco_executor_sc:
                            if not (
                                    exists(endereco_executor_sc, "endereço prestador", "", f'EXECUTOR={result[0]["id"]}') or
                                    exists(endereco_executor_sc, "endereço prestador", " ", f'EXECUTOR={result[0]["id"]}')
                                ):
                                return 'Endereço do prestador não encontrado ou inválido'

                return None                   

            def get_peg() -> str:

                locator = safe_locator(page,"//span[@data-field='PEG']//xmp")
                if locator.count() == 0:
                    return None
                    
                return locator.first.inner_text().strip()
                
            def get_error_message() -> str:

                locator = safe_locator(page,"xpath=//div[contains(@class, 'message-panel')]//span[last()]")
                if locator.count() == 0:
                    return None
                
                error_message = str(re.sub('\s+',' ',locator.first.inner_text()).strip()) 

                if "porém a gravação do registro será permitida" in error_message:
                    logging.info(error_message)
                    return None 
                
                if "O Beneficiário Titular do PEG é também o titular da família do beneficiário desta guia!" in error_message:
                    logging.info(error_message)
                    return None                 

                return error_message            
    
            @timeit
            @handle_exceptions(default_return=False)
            @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)            
            def delete_peg(peg) -> bool:

                if not global_parameters["benner.delete_peg"] or not peg:
                    return False

                logging.info(f'Deletando o peg {peg}...')

                if not is_element_ready(page,'#breadcrumbUpdatePanel'):
                    logging.error(f'breadcrumb não encontrado')            
                    return False

                with page.expect_navigation():
                    safe_locator(page,'#breadcrumbUpdatePanel a', has_text=peg.replace(".","")).click()

                page.evaluate("""
                    (client_id) => {
                        __doPostBack('ctl00$Main$WDG_V_SAM_PEG_'+client_id+'_PRINCIPAL','CMD_EXCLUIRPEG');
                    }
                """, client_id)   

                wait_for_load_state(page)      # javascript:__doPostBack('ctl00$Main$WDG_V_SAM_PEG_1465_PRINCIPAL$RQTrue','')

                if not is_element_ready(page,'input[type="button"][value="Continuar"]'):
                    logging.error(f'botão "Continuar" não encontrado')
                    return False          
                
                with page.expect_navigation():
                    safe_locator(page,'input[type="button"][value="Continuar"]').click()

                wait_for_load_state(page)

                message = None
                xpath = '#ctl00_TopBarMainContent_globalMessagePanel_message'

                if is_element_ready(page, xpath):
                    message = safe_locator(page, xpath).inner_text()
                    logging.info(message)                 
                
                return message == "Processo enviado para execução no servidor! Peg sendo excluído!"

            nonlocal executor_sc, endereco_executor_sc
            error_message = None
            guide_number = None
            peg = None
            client_id = None
            
            try:

                logging.info(f'Criando PEG...')    

                # preenche os dados do PEG
                guide_number = None
                peg = None
                client_id = None
                steps = [valid_required_fields, fill_peg, fill_guide]

                for step in steps:
                    error_message = step()
                    if error_message:
                        logging.error(f"Erro na etapa '{step.__name__}': {error_message}")
                        break

            except Exception as e:

                page.screenshot(path=f'{log}\\screenshot_{today}_{today_.strftime("%H%M%S")}.png', full_page=True)
                error_message = str(e)
                logging.error(traceback.format_exc())
            
            finally:

                # atualizar Excel
                logging.info("atualizando excel...")

                row[colunas.get("PEG")].value = None if error_message else peg
                row[colunas.get("comment")].value = error_message
                row[colunas.get("guide_number")].value = guide_number

                wb.save(worksheet)     

                # exclui o peg neste caso
                if error_message and peg:
                    delete_peg(peg)             

                # # Selecione a filial
                # with page.expect_navigation():
                #     safe_locator(page,'#breadcrumbUpdatePanel a', has_text=global_parameters["benner.filial"]).click()

            return not error_message

        @timeit
        @handle_exceptions(default_return=False)
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def login() -> bool:

            with page.expect_navigation():
                page.goto(global_parameters[f'benner.url_{global_parameters["env"]}'])      

            safe_locator(page,"input[name='wesLogin$loginWes$UserName']").first.fill(config["benner"]["username"])
            safe_locator(page,"input[name='wesLogin$loginWes$Password']").first.fill(config["benner"]["password"])
            
            with page.expect_navigation():
                safe_locator(page,"//*[@id='LoginButton']").click()

            logging.info(f'Login benner com sucesso {config["benner"]["username"]}...')

            # Competências de PEG
            with page.expect_navigation():
                page.goto(global_parameters[f'benner.competenciasdepeg_{global_parameters["env"]}'])

            # Selecione a competência
            with page.expect_navigation():
                safe_locator(page,f'//a[text()={global_parameters["benner.competencia"]}]').click()

            # cliar em FILIAIS
            safe_locator(page,f'//div/span[@title="FILIAIS"]').first.click()
            
            # Selecione a filial
            with page.expect_navigation():
                safe_locator(page,f'//a[text()="{global_parameters["benner.filial"]}"]').click()  

            logging.info(f'Seleção da filial {global_parameters["benner.filial"]} com sucesso')

            return True          

        @timeit
        @handle_exceptions(default_return=None)
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def goto_filial() -> tuple[dict, str]:

            # Competências de PEG
            with page.expect_navigation():
                page.goto(global_parameters[f'benner.competenciasdepeg_{global_parameters["env"]}'])

            # Selecione a competência
            with page.expect_navigation():
                safe_locator(page,f'//a[text()={global_parameters["benner.competencia"]}]').click()

            # cliar em FILIAIS
            safe_locator(page,f'//div/span[@title="FILIAIS"]').first.click()
            
            # Selecione a filial
            with page.expect_navigation():
                safe_locator(page,f'//a[text()="{global_parameters["benner.filial"]}"]').click()  

            # 1. Captura o form e a URL correta
            form_data, action_url = capture_aspnet_form(page)       

            logging.info(f'Seleção da filial {global_parameters["benner.filial"]} com sucesso')

            return form_data, action_url               
        
        logging.info(f'Acessando o site...')
        logging.info(global_parameters[f'benner.url_{global_parameters["env"]}'])
        logging.info(f'Usuário {config["benner"]["username"]}')

        browser = playwright.chromium.launch(
            headless=bool(global_parameters["benner.headless"]), 
            args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"]
        )

        if bool(global_parameters["benner.record_video"]):
            context = browser.new_context(record_video_dir=log)
        else:
            context = browser.new_context()

        page = context.new_page()

        # Define o timeout padrão (em milissegundos)
        page.set_default_timeout(int(global_parameters["timeout"]))
        page.set_default_navigation_timeout(int(global_parameters["timeout"]))        
        
        page.route("**/*", lambda route: route.abort() if route.request.resource_type in ["image", "font"] else route.continue_()) #, "stylesheet", "font"
         
        #
        wb = safely_load_workbook(worksheet)

        if not wb:
            return 0, 0 

        ws = wb.active  # Pega a primeira planilha ativa    

        # Criar um dicionário para mapear os nomes das colunas aos seus índices
        #colunas = {cell.value: cell.column for cell in ws[1]}  # Pega a linha 1 como cabeçalho
        protocols_count = 0
        protocols_total = ws.max_row -1
        executor_sc = None
        endereco_executor_sc = None

        if not login():
            return protocols_total, 0 
        
        form_data, action_url = goto_filial()

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

            protocols_count += 1

            if protocols_count % int(global_parameters["benner.bloco_login"]) == 0:
                if not login():
                    logging.error(f'Criação dos PEGs interrompida no protocolo {protocols_count}/{protocols_total}')
                    break            
                form_data, action_url = goto_filial()
        
            logging.info(f'Processando protocolo {row[colunas.get("protocol_id")].value} {protocols_count}/{protocols_total}...')
            
            if row[colunas.get("comment")].value:
                logging.warning(row[colunas.get("comment")].value)   

            if not row[colunas.get("PEG")].value:
                protocols_success += (1 if add_peg() else 0)
            else:
                logging.info(f'PEG {row[colunas.get("PEG")].value} criado anteriormente')     

        # 
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
           
            if not row[colunas.get("comment")].value and not row[colunas.get("PEG")].value:
                row[colunas.get("comment")].value = f'Processamento interrompido'  
                logging.warning(f'Processamento interrompido do protocolo {row[colunas.get("protocol_id")].value}')

        if bool(global_parameters["benner.record_video"]):
            logging.info(f'vídeo gravado em {page.video.path()}')

        context.close()

        wb.save(worksheet)
    
    return protocols_total, protocols_success   
        
@timeit
def update_protocols_mobile_saude(worksheet) -> list:
    
    logging.info(f'Atualizando protocolos no mobilesaude...')

    protocols_total = 0
    protocols_count = 0
    protocols_success = 0       

    with sync_playwright() as playwright:

        def get_error_message() -> str:

            error_message = None
            locator = safe_locator(page,"div.toast.toast-error div.toast-message")

            if locator.count():
                error_message = locator.first.inner_text().strip()

            locator = safe_locator(page,"div.msg-box.success")

            if locator.count():
                error_message = locator.first.inner_text().strip()

            if error_message and (error_message == "SUCCESS" or error_message == "Atualização de reembolso atualizado com sucesso!"):
                return None

            return error_message

        @timeit
        @handle_exceptions()
        def update_protocol(change_status):

            error_message = None 

            url = global_parameters["mobilesaude.reembolso"] %(row[colunas.get("refund_id")].value)
            buffer = ""

            logging.info(f'Abrindo {url}')

            # abre página de edição do protocolo
            with page.expect_navigation():
                response = page.goto(url)

            if response and response.status != HTTPStatus.OK:
                return f'Não foi possível abrir a página de edição do protocolo {row[colunas.get("refund_id")].value}'
            
            observacao_interna = safe_locator(page,f'xpath=//textarea[@name="despesa[{row[colunas.get("refund_qty")].value or 1}][observacao_interna]"]').first

            PEG = match(r'\d{2}\.\d{3}\.\d{3}', observacao_interna.input_value() or "")   
            PEG = safe_locator(page, '//input[@name="numero_lote"]' ).inner_text() or PEG
            
            assigned = None
            locator = safe_locator(page, '//p[text()="Atribuição: "]/strong')
            if locator.count():
                assigned = locator.inner_text()      
                           
            status_id = safe_locator(page,'//select[@name="id_status_reembolso"]').input_value()
            status_desc = safe_locator(page,'//select[@name="id_status_reembolso"]/option[@selected]').text_content()

            if PEG or status_id != str(global_parameters[f'mobilesaude.status_filtro_{global_parameters["env"]}']):
                
                row[colunas.get("assigned")].value = assigned
                row[colunas.get("complement")].value = f'Protocolo já estava atualizado para {status_desc}{" por " + assigned if assigned else ""}{" com o PEG " + PEG if PEG else ""}'

                logging.info(row[colunas.get("complement")].value)
                return None

            if not assign(global_parameters["mobilesaude.bot_user_id"], global_parameters["mobilesaude.bot_user_name"]):
                return f'Não foi possível atribuir {row[colunas.get("protocol_id")].value} a {global_parameters["mobilesaude.bot_user_name"]}'

            with page.expect_navigation():
                page.reload()

            row[colunas.get("assigned")].value = global_parameters["mobilesaude.bot_user_name"]   
            
            if row[colunas.get("PEG")].value:
                buffer += f'PEG:{row[colunas.get("PEG")].value}'

            # se existir mensagem de erro, checar se deve ser gravada na observação
            if row[colunas.get("comment")].value and global_parameters["mobilesaude.update_protocol_with_error"]:
                
                if buffer:
                    buffer += '\n'

                buffer += f'{row[colunas.get("comment")].value}'

            if buffer:
                observacao_interna.fill(f'{buffer}\n{observacao_interna.input_value()}')

            # atualizar o status do protocolo para em análise se tem PEG e não tem comentário de erro
            if change_status:
                page.evaluate('''
                    (params) => {
                        $(params.selector).val(params.value).trigger("change")
                    }
                ''', {"selector": "div.form-group:has(label:contains('STATUS REEMBOLSO')) select", "value": str(global_parameters[f'mobilesaude.status_final_protocolo_{global_parameters["env"]}'])})  
            
            # Atualiza os ststus
            #row[colunas.get("status_id")].value = safe_locator(page,'//select[@name="id_status_reembolso"]').input_value()
            #row[colunas.get("status_desc")].value = safe_locator(page,'//select[@name="id_status_reembolso"]/option[@selected]').text_content()

            if not is_element_ready(page, 'role=button[name="Atualizar Solicitação e Sair"]', check="enabled"):
                return f'botão "Atualizar Solicitação e Sair" não está habilitado'               

            safe_locator(page, 'role=button[name="Atualizar Solicitação e Sair"]').click()

            if change_status:

                if not is_element_ready(page, "#atribuir-status-reembolso-manual"):
                    return f'Janela de seleção do status do protocolo não está aberta {get_error_message()}'             
                
                # Localiza e clica no botão da linha correspondente ao status
                botao = safe_locator(page,f'xpath=//tr[td[contains(text(), "Em Análise")]]//button')
                if not botao.count():
                    return f'botão "Em Análise" não encontrado'             
                
                with page.expect_navigation():
                    botao.click()

            # 
            error_message = get_error_message()
            if error_message:
                return error_message     

            logging.info(f'Protocolo {row[colunas.get("protocol_id")].value} atualizado com sucesso')        

            return None

        @timeit
        @handle_exceptions(default_return=False)
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def assign(owner_id, owner_name) -> bool:

            if not bool(global_parameters["mobilesaude.assign"]):
                return False            
            
            refund_id = row[colunas.get("refund_id")].value
            protocol_id = row[colunas.get("protocol_id")].value

            payload = {
                "id_reembolso": str(refund_id),
                "id_ms_usuario": str(owner_id)
            }

            # Usar o contexto com sessão ativa
            request_context = context.request

            response = request_context.post(
                global_parameters["mobilesaude.atribuir"],
                multipart=payload,
                ignore_https_errors=True
            )                

            # check if the request was successful
            if response.status != HTTPStatus.OK:
                logging.error(f"Falha na atribuição do protocolo {protocol_id} a {owner_name}: {response.status} {response.text()}")  
                return False

            logging.info(f'Protocolo {protocol_id} atribuido com sucesso ao usuário {owner_name}...')

            return True   
        
        @timeit
        @handle_exceptions(default_return=False)
        @retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
        def login() -> bool:

            with page.expect_navigation():
                page.goto(global_parameters["mobilesaude.url"])      

            safe_locator(page,"input[name='email']").first.fill(config["mobilesaude"]["username"])
            safe_locator(page,"input[name='senha']").first.fill(config["mobilesaude"]["password"])

            with page.expect_navigation():
                safe_locator(page,"button:has-text('Acessar')").click()

            logging.info(f'Login com sucesso no Mobilesaude {config["mobilesaude"]["username"]}...')

            # SAUDE AMS
            logging.info(f'Navegando para SAUDE AMS...')

            with page.expect_navigation():
                response = page.goto(global_parameters[f'mobilesaude.saude_ams_{global_parameters["env"]}'])            

            # check if the request was successful
            if response and response.status != HTTPStatus.OK:      
                logging.error(f'Não foi possível acessar o módulo Saúde AMS')   
                return False
            
            logging.info(f'Navegando para SAUDE AMS... OK')

            return True   
        
        logging.info(f'Acessando o site...')
        logging.info(f'{global_parameters["mobilesaude.url"]}')
        logging.info(f'Usuário {config["mobilesaude"]["username"]}')

        browser = playwright.chromium.launch(
            headless=bool(global_parameters["mobilesaude.headless"]), 
            args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"]
        )

        if bool(global_parameters["mobilesaude.record_video"]):
            context = browser.new_context(record_video_dir=log)
        else:
            context = browser.new_context()

        page = context.new_page()

        # Define o timeout padrão (em milissegundos)
        page.set_default_timeout(int(global_parameters["timeout"]))
        page.set_default_navigation_timeout(int(global_parameters["timeout"]))

        # PNG transparente 1x1 codificado em base64
        imagem_base64 = (
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8"
            "/w8AAgMBAQEAkBkAAAAASUVORK5CYII="
        )
        imagem_substituta = base64.b64decode(imagem_base64)

        def intercept_request(route, request):
            if request.resource_type in ["image"]:
                route.fulfill(
                    status=200,
                    content_type="image/png",
                    body=imagem_substituta
                )
            elif request.resource_type in ["font"]:
                route.abort()
            else:
                route.continue_()

        page.route("**/*", intercept_request)

        # carregar o excel
        wb = safely_load_workbook(worksheet)

        if not wb:
            return 0, 0      
        
        ws = wb.active  # Pega a primeira planilha ativa    

        # Criar um dicionário para mapear os nomes das colunas aos seus índices
        #colunas = {cell.value: cell.column for cell in ws[1]}  # Pega a linha 1 como cabeçalho
        protocols_count = 0
        protocols_total = ws.max_row -1

        if not login():
            return protocols_total, 0         
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            
            protocols_count += 1

            logging.info(f'Consultando o protocolo {row[colunas.get("protocol_id")].value} {protocols_count}/{protocols_total}...')

            change_status = (not row[colunas.get("comment")].value) and (row[colunas.get("PEG")].value)

            if change_status or global_parameters["mobilesaude.update_protocol_with_error"]:

                error_message = update_protocol(change_status)
                if error_message:
                    if isinstance(error_message, Exception):
                        error_message = str(error_message)
                    logging.error(error_message)

                # se não tinham comentário de erro, verifica se agora tem
                if not row[colunas.get("comment")].value and error_message: 
                    row[colunas.get("comment")].value = error_message
                
                # se tem comentário de erro, checa se precisa atribuir para usuário humano 
                if row[colunas.get("comment")].value and global_parameters["mobilesaude.assign_alternative"]:
                    if assign(global_parameters["mobilesaude.alternative_user_id"], global_parameters["mobilesaude.alternative_user_name"]):
                        row[colunas.get("assigned")].value = global_parameters["mobilesaude.alternative_user_name"]   

                # se terminou o processo sem comentário de erro, setar "OK"
                if not row[colunas.get("comment")].value:
                    row[colunas.get("comment")].value = "OK"

                wb.save(worksheet) # atualiza o Excel        

            else:
                logging.warning(f'Nada para atualizar no prococolo {row[colunas.get("protocol_id")].value}...')

            protocols_success += 1 if row[colunas.get("comment")].value == "OK" else 0

        wb.save(worksheet)           
        
        if bool(global_parameters["mobilesaude.record_video"]):
            logging.info(f'vídeo gravado em {page.video.path()}')

        context.close()
    
    return protocols_total, protocols_success 
    
@timeit
def reload_old_files():

    def find_xlsx_files(base_directory):
        xlsx_files = []

        for root, dirs, files in os.walk(base_directory):
            for file in files:
                if file.lower().startswith('protocolos') and file.lower().endswith('.xlsx'):
                    full_path = os.path.join(root, file)
                    xlsx_files.append(full_path)

        return xlsx_files

    logging.info(f'Recarregando arquivos de processamentos anteriores...')

    worksheets = find_xlsx_files(f'{path}\\log')
    worksheets_count = 0
    worksheets_total = len(worksheets)

    column_names = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="labels", header=None).iloc[:, 0].tolist()
    column_ids = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="ids", header=None).iloc[:, 0].tolist()

    db = LMDBWrapper()

    logging.info(db.usage_stats())

    for worksheet in worksheets:

        worksheets_count += 1         

        logging.info(f'Reprocessando planilha {os.path.basename(worksheet)} {worksheets_count}/{worksheets_total}...')

        result = db.count_where(lambda d: d["file_id"] == os.path.basename(worksheet))     
        
        #
        wb = safely_load_workbook(worksheet, read_only=True)
        if not wb:
            continue
        
        ws = wb.active
        protocols_total = ws.max_row -1

        if result == protocols_total:
            continue

        colunas = {cell.value: cell.column for cell in ws[1]}
        new_colunas = {}

        for chave, valor in colunas.items():
            if chave in column_names:
                index = column_names.index(chave)
                nova_chave = column_ids[index]
            else:
                nova_chave = chave  # mantém original se não encontrar
            new_colunas[nova_chave] = valor - 1 

        colunas = new_colunas
        protocols_count = 0

        db.delete_where(lambda doc: doc["file_id"] == os.path.basename(worksheet))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

            protocols_count += 1         

            logging.info(f'Incluindo protocolo {row[colunas.get("protocol_id")].value} no BD {protocols_count}/{protocols_total}...')

            record = {
                'file_id': os.path.basename(worksheet),
                'env': row[colunas.get("env")].value if colunas.get("env") else ("prd" if int(row[colunas.get("refund_id")].value) > 999999 else "hml"),
                'refund_id': row[colunas.get("refund_id")].value,
                'protocol_id': row[colunas.get("protocol_id")].value,
                'protocol_date': row[colunas.get("protocol_date")].value,
                'status_id': row[colunas.get("status_id")].value,
                'status_desc': row[colunas.get("status_desc")].value,
                'refund_type': row[colunas.get("refund_type")].value,
                'refund_qty': row[colunas.get("refund_qty")].value,
                'refund_value': row[colunas.get("refund_value")].value,
                'card': row[colunas.get("card")].value,
                'user': row[colunas.get("user")].value if colunas.get("user") else None,
                'holder_name': row[colunas.get("holder_name")].value,
                'holder_cpf': row[colunas.get("holder_cpf")].value,
                'phone_number': row[colunas.get("phone_number")].value,
                'plan': row[colunas.get("plan")].value,
                'payment_day': row[colunas.get("payment_day")].value,
                'payment_type': row[colunas.get("payment_type")].value,
                'lot': row[colunas.get("lot")].value,
                'expense_id': row[colunas.get("expense_id")].value,
                'expense_status': row[colunas.get("expense_status")].value,
                'supplier_id': row[colunas.get("supplier_id")].value,
                'supplier_name': row[colunas.get("supplier_name")].value,
                'supplier_state': row[colunas.get("supplier_state")].value,
                'supplier_city': row[colunas.get("supplier_city")].value,
                'expense_date': row[colunas.get("expense_date")].value,
                'expense_nf': row[colunas.get("expense_nf")].value,
                'guide_number': row[colunas.get("guide_number")].value if colunas.get("guide_number") else None,
                'PEG': row[colunas.get("PEG")].value,
                'notes': row[colunas.get("notes")].value,
                'assigned': row[colunas.get("assigned")].value,
                'comment': row[colunas.get("comment")].value,
                'complement': row[colunas.get("complement")].value if colunas.get("complement") else None,
            }

            db.insert(record)

    logging.info(f'Recarregando arquivos de processamentos anteriores... OK')

    return None

@retry(retry_on_exception=retry_if_stop_exception, 
       wait_fixed=int(config["restart"]["wait_fixed"]), 
       stop_max_attempt_number=int(config["restart"]["stop_max_attempt_number"]))  
@timeout(int(config["restart"]["max_execution_time"]), timeout_exception=StopIteration, dec_poll_subprocess=2)
@timeit
def run() -> int: 

    def send_emails(sumary):
        
        if not bool(config["smtp"]["enabled"]):
            return

        with open(config["smtp"]["template"], 'r', encoding=config["commons"]["encoding"]) as template:
            
            html = template.read().format(
                "A tarefa de lançamento de PEG foi realizada com sucesso",
                "A tarefa de lançamento de PEG foi realizada com sucesso. Todos os detalhes do processamento estão no log em anexo.",
                "<p style=\"font-family:'Courier New'\">" + "<br/>".join(sumary) + "</p>",
                "",
                "lançamento PEG"
                )
        
        logging.info("Enviando e-mail...")
        logging.info(config["emails"]["success"] + "," + str(global_parameters["emails_success"]))
        logging.shutdown()

        sendemail_postmarkapp(config["smtp"]["host"], 
                config["smtp"]["port"], 
                config["smtp"]["username"], 
                config["smtp"]["password"], 
                config["smtp"]["headers"], 
                f'{config["smtp"]["subject"]} - AGENDA {todayFormatted} {today_.strftime("%H:%M")}', 
                config["smtp"]["from"], 
                config["emails"]["success"] + "," + str(global_parameters["emails_success"]), 
                html,
                [config["smtp"]["logo"]],
                [f"{log}/{arquivo}" for arquivo in os.listdir(log)]) 
        
    try:

        # Criar um contexto SSL que ignora a verificação de certificado
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE

        os.environ['SSL_CERT_FILE'] = certifi.where()            

        config = getConfig()

        bot_base()

        banner()

        logging.info('Iniciando a execução da automação...')

        for key, value in global_parameters.items(): logging.info(f"{key}: {value}")            

        delete_temp_files()

        # if not os.path.exists('db.lmdb'):
        #     db = LMDBWrapper()
        #     tinydb = TinyDB('db.json')
        #     all_data = tinydb.all()

        #     for item in all_data:
        #         if "file_id" in item:
        #             db.insert(item)   
                
        reload_old_files()

        logging.info(f'log {log}')

        protocols_total = 0
        protocols_success = 0        

        if bool(global_parameters["mobilesaude.read_protocols_mobile_saude"]):
            worksheet = read_protocols_mobile_saude() 
        else:
            worksheet = global_parameters["mobilesaude.worksheet"]
                
        if bool(global_parameters["benner.create_peg_benner"]) and worksheet:
            protocols_total, protocols_success = create_peg_benner(worksheet)

        if bool(global_parameters["mobilesaude.update_protocols_mobile_saude"]) and worksheet:
            protocols_total, protocols_success = update_protocols_mobile_saude(worksheet)

        export_backlog()

        endTime = timer()
        tempo_medio = ((endTime-startTime)/protocols_total) if protocols_total else 0
        saving = float(global_parameters["saving"])

        sumary = [
            "Protocolos processados..............: " + str(protocols_total),
            "Protocolos processados com sucesso..: " + str(protocols_success),
            "Protocolos processados com erro.....: " + str(protocols_total-protocols_success),
            "Tempo médio por protocolo...........: " + humanize.precisedelta(tempo_medio),
            "Tempo de processamento..............: " + humanize.precisedelta(timedelta(seconds=endTime-startTime)),
            "Tempo economizado...................: " + humanize.precisedelta(protocols_total * (saving - tempo_medio))
        ]

        for x in sumary: logging.info(x)

        finish()

        logging.info("#### Finalizado ####")     

        send_emails(sumary)

        return 0

    except:
        show_exception_and_exit(*sys.exc_info())    