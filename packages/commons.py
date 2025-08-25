import pandas as pd
import os
import sys
import re
import requests
import json
import logging
import psutil
from datetime import datetime
from unicodedata import normalize
from unicodedata import combining
from retrying import retry
import smtplib
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.utils import formataddr
from os.path import basename
import win32com.client
import sys
import subprocess
import time
import pythoncom
import win32ui
from subprocess import check_output
from xml.dom.minidom import *
from datetime import datetime, timedelta
import paramiko
from functools import wraps, reduce
import time
import tempfile
from playwright.sync_api import TimeoutError, Page
import functools
from urllib.parse import urljoin
import shutil
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from playwright.sync_api import Locator
from typing import List, Callable
import lmdb
import uuid
from packages.config import config
from packages.global_parameters import global_parameters
from pathlib import Path
from datetime import datetime
from requests import Session, session
from packages.bot_base import log
from packages.bot_base import today_
from packages.bot_base import today
from packages.bot_base import path
from http import HTTPStatus
from collections import Counter
from requests import Session

xmlParser = Document()

def banner():
    logging.info("   ____       _      ____       _     ")
    logging.info("  |  _ \     / \    / ___|     / \    ")
    logging.info("  | |_) |   / _ \   \___ \    / _ \   ")
    logging.info("  |  __/   / ___ \   ___) |  / ___ \  ")
    logging.info("  |_|     /_/   \_\ |____/  /_/   \_\ ")    
    logging.info("                                      ") 

# def safe(f):
#     """return a new function that's the same as f,
#     except that it outputs infinity whenever f produces an error"""
#     def safe_f(*args, **kwargs):
#         try:
#              return f(*args, **kwargs)
#         except:
#              return None
#     return safe_f                   

# Função para normalizar traços
def normalize_hyphens(s):
    # Substitui diferentes tipos de traços por '-'
    return re.sub(r'[\u2010\u2011\u2012\u2013\u2014\u2015]', '-', s)

# Função para processar o dicionário de maneira genérica
def normalize_dict_hyphens(data):
    if isinstance(data, dict):
        # Se for um dicionário, processa recursivamente as chaves e valores
        return {key: normalize_dict_hyphens(value) for key, value in data.items()}
    elif isinstance(data, list):
        # Se for uma lista, processa cada elemento
        return [normalize_dict_hyphens(item) for item in data]
    elif isinstance(data, str):
        # Se for uma string, aplica a normalização de traços
        return normalize_hyphens(data)
    else:
        # Se não for string, lista ou dicionário, retorna o valor original
        return data
    
def get_parameters(report_name = None):
    path = './data' if not report_name else f'./data/{report_name}'
    return __get_parameters(f'{path}/parameters.xlsx')     

def __get_parameters(xlsx="./data/parameters.xlsx", sheet_name="values", key="key", value="value"):
    df = pd.read_excel(xlsx, engine='openpyxl', sheet_name=sheet_name)
    return normalize_dict_hyphens(dict(zip(df[key], df[value])))     

def timeit(func):
    @wraps(func)
    def timeit_wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        result = func(*args, **kwargs)
        end_time = time.perf_counter()
        total_time = end_time - start_time
        # first item in the args, ie `args[0]` is `self`
        logging.info(f'Function {func.__name__} Took {total_time:.4f} seconds')
        return result
    return timeit_wrapper

def sftpOpen(host, port = 22, username = None, password = None):
    # create ssh client 
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_client.connect(hostname=host,port=port,username=username,password=password)

    return ssh_client, ssh_client.open_sftp()

def sftpClose(client, session):
    # close the connection
    session.close()
    client.close()    

def checkIfProcessRunning(processName):
    '''
    Check if there is any running process that contains the given name processName.
    '''
    #Iterate over the all the running process
    for proc in psutil.process_iter():
        try:
            # Check if process name contains the given name string.
            if processName.lower() in proc.name().lower():
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False

def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    processes = []
    for process in check_output(call).splitlines()[3:]:
        process = process.decode()
        processes.append(process.split())
    return processes

# def process_exists(process_name):
#     progs = str(subprocess.check_output('tasklist'))

#     if process_name in progs:
#         return True
#     else:
#         return False

def WindowExists(classname):
    try:
        win32ui.FindWindow(classname, None)
    except win32ui.error:
        return False
    else:
        return True
    
def is_process_running(process_name):
    cmd = 'tasklist /fi "imagename eq {}"'.format(process_name)
    output = subprocess.check_output(cmd, shell=True).decode()
    if process_name.lower() in output.lower():
        return True
    else:
        return False

def DrowWaitCursor(counter):
    if counter % 4 == 0:
        print("/",end = "")
    elif counter % 4 == 1:
        print("-",end = "")
    elif counter % 4 == 2:
        print("\\",end = "")
    elif counter % 4 == 3:
        print("|",end = "")
    sys.stdout.flush()
    sys.stdout.write('\b') 

def spinner():
    chars = '|/-\\'
    for i in range(30):
        for c in chars:
            sys.stdout.write(' %3d :: %s\r' % (i, c))
            sys.stdout.flush()
            time.sleep(0.2)

def proc_v(column, search, df):

    result = df.loc[ df[column] == search ]

    if len(result) > 0:
        return result.squeeze()
    else:
        return None
    
def saplogin(path, sid, client, user, password, language, login_max_attempts):

    pythoncom.CoUninitialize()
    pythoncom.CoInitialize()

    path = f'{path} -system={sid} -client={client} -user={user} -pw={password} -language={language}'
    
    #path + " -system=" + sid + " -client=" + client + " -user=" + user + " -pw=" + password + " -language=" + language

    subprocess.Popen(path)
    attempts = 0
    time.sleep(1)
    shell = win32com.client.Dispatch("WScript.Shell")
    time.sleep(1)

    logging.info("Wait for SAP Easy Access")
    while not shell.AppActivate("SAP Easy Access") and attempts <= login_max_attempts:
        time.sleep(1)
        attempts = attempts + 1

        if attempts > 10:
            connection = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0)

            if connection:
                Wnd1 = findById(connection,"ses[0]/wnd[1]")

                if Wnd1: 
                    logging.warning(f'Janela inesperada identificada {Wnd1.text}')
                    Wnd1.close()
                            
        DrowWaitCursor(attempts)
        #spinner()
        
    if attempts > login_max_attempts:	
       raise StopIteration("SAP Easy Access error on logon")

    SapGuiAuto = win32com.client.GetObject('SAPGUI')
    if not type(SapGuiAuto) == win32com.client.CDispatch:
        return

    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32com.client.CDispatch:
        SapGuiAuto = None
        return
    connection = application.Children(0) # application.OpenConnection("ConnectionName", True)

    if not type(connection) == win32com.client.CDispatch:
        application = None
        SapGuiAuto = None
        return

    session = connection.Children(0)
    if not type(session) == win32com.client.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    
    return session

def getCountVerticalScrollDown(session, id):

    vertScrollBar = session.findById(id).VerticalScrollbar
    vertScrollBar.Position = vertScrollBar.Minimum

    if (vertScrollBar.Range > 0):
        cntScrollDown = (0 - int(0 - (vertScrollBar.Maximum / vertScrollBar.Range)))
    else:
        cntScrollDown = 1

    return cntScrollDown
	
# Starting the transaction
def startTransaction(session, transaction_name):
	#logging.info("Transaction " & transaction_name & "... ") 
	session.findById("wnd[0]/tbar[0]/okcd").Text = transaction_name
	pressEnter(session.findById("wnd[0]"))
	#time.sleep(5)

#https://experience.sap.com/files/guidelines/References/nv_fkeys_ref2_e.htm

# Pressing the "Enter"
def pressEnter(window):
	window.sendVKey(0)

# Pressing the "Double Click"
def pressDblClick(window):
	window.sendVKey(2)    

# Pressing the "F2"
def pressF2(window):
	window.sendVKey(2)
      
# Pressing the "F3"
def pressF3(window):
	window.sendVKey(3)
     
# Pressing the "F5"
def pressF5(window):
	window.sendVKey(5)         

# Pressing the "F8"
def pressF8(window):
	window.sendVKey(8)    
     
# Pressing the "F9"
def pressF9(window):
	window.sendVKey(9)        

# Pressing the "F7"
def pressF7(window):
	window.sendVKey(7)        

# Pressing the "F12"
def pressF12(window):
	window.sendVKey(12)        

# Pressing the "Page Down"
def pressPageDown(window):
	window.sendVKey(82)

# Pressing the "Shift-F2"
def pressShiftF2(window):
	window.sendVKey(14)     

# Pressing the "Shift-F4"
def pressShiftF4(window):
	window.sendVKey(16)    

# Pressing the "Ctrl-Shift-F2"
def pressCtrlShiftF2(window):
	window.sendVKey(38)    

# Pressing the "Ctrl-F5"
def pressCtrlF5(window):
	window.sendVKey(29)    

# Pressing the "Ctrl-S"
def pressCtrlS(window):
	window.sendVKey(11) # or F11      

# Pressing the "Shift-F5"
def pressShiftF5(window):
	window.sendVKey(17)   

# Pressing the "Ctrl-P"
def pressCtrlP(window):
	window.sendVKey(86)      

# Pressing the "Shift-F12"
def pressShiftF12(window):
	window.sendVKey(24)    

# Pressing the "Ctrl-Shift-F8"
def pressCtrlShiftF8(window):
	window.sendVKey(44)  

# Pressing the "Ctrl-Shift-F9"
def pressCtrlShiftF9(window):
	window.sendVKey(45)  

# Pressing the "Ctrl-Shift-F12"
def pressCtrlShiftF12(window):
	window.sendVKey(48)                     

#Back to initial screen
def backToInitialScreen(session):
    session.findById("wnd[0]/tbar[0]/okcd").Text = "/ns000"
    session.findById("wnd[0]/tbar[0]/btn[0]").Press()

def logoff(session):
	
	session.findById("wnd[0]").close()
    
	if (session.findById("wnd[1]/usr/btnSPOP-OPTION1")):
		session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()

def close_excel():
    for proc in psutil.process_iter():
        if proc.name() == "excel.exe" or proc.name() == "EXCEL.EXE":
            proc.kill()

def close_sap():
    for proc in psutil.process_iter():
        if proc.name() == "saplogon.exe":
            proc.kill()            

def change_date_format(dt):
    return re.sub(r'(\d{4})-(\d{1,2})-(\d{1,2})', '\\3-\\2-\\1', dt)

def autosize_excel_columns(worksheet, df):
    autosize_excel_columns_df(worksheet, df.index.to_frame())
    autosize_excel_columns_df(worksheet, df, offset=df.index.nlevels)


def autosize_excel_columns_df(worksheet, df, offset=0):
    for idx, col in enumerate(df):
        series = df[col]
        max_len = max((
            series.astype(str).map(len).max(),
            len(str(series.name))
        )) + 1
        worksheet.set_column(idx+offset, idx+offset, max_len)


def get_month_year(_month, _year):

    if int(_month) > 9:
        return _year + '-' + _month + '-01'
    else:
        return _year + '-0' + _month + '-01'


def custom_styles(val):
    # price column styles
    if val.name == 'price':
        styles = []
        # red prices with 0
        for i in val:
            styles.append('color: %s' % ('red' if i == 0 else 'black'))
        return styles
    # other columns will be yellow
    return ['background-color: yellow'] * len(val)


def calc_interval(date1, date2):

    if date1 and date2:
        return pd.to_datetime(date1) - pd.to_datetime(date2)
    elif date1 == None or date1 == '':
        return datetime.now() - pd.to_datetime(date2)
    else:
        return ''


def excel_column_number(name):
    """Excel-style column name to number, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    n = 0
    for c in name:
        n = n * 26 + 1 + ord(c) - ord('A')
    return n


def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(r + ord('A')) + name
    return name


def auto_width_columns(df, worksheet):
    for i, col in enumerate(df.columns):
        column_len = max(df[col].astype(str).str.len().max(), len(col) + 2)
        worksheet.set_column(i, i, column_len)

def in_between(minv, val, maxv) -> bool:
    return val == min(maxv, max(minv, val))

def set_decimal_point(value):
    buffer = re.sub('\.', '', str(value))
    buffer = re.sub(',', '.', buffer)
    return buffer

def adjust_negative_sign(value):

    buffer = str(value)
    
    if buffer.find("-") > -1:
         buffer = re.sub('-', '', str(value))
         buffer = f'-{buffer}'

    return buffer

def remove_accents(s):
    nkfd_form = normalize('NFKD', s)
    return u''.join([c for c in nkfd_form if not combining(c)])


def diff_month(date1, date2):
    return (date1.year - date2.year) * 12 + date1.month - date2.month

def finish():
    logging.info("    ____     ")
    logging.info("   [____]    ")
    logging.info(" |=]()()[=|  ")
    logging.info(" __\_==_/__  ")
    logging.info("|__|    |__| ")
    logging.info(" |_|_/\_|_|  ")
    logging.info(" | | __ | |  ")
    logging.info(" |_|[  ]|_|  ")
    logging.info(" \_|_||_|_/  ")
    logging.info("   |_||_|    ")
    logging.info("  _|_||_|_   ")
    logging.info(" |___||___|  ")
    logging.info("             ")

def retry_if_connection_error(exception):
    logging.error(exception)
    return isinstance(exception, ConnectionError)

def retry_if_stop_exception(exception):
    logging.info("Executando retry_if_stop_exception")
    logging.error(exception)
    return isinstance(exception, StopIteration)

def retry_if_getaddrinfo_failed(exception):
    logging.error(exception)

    if hasattr(exception, 'errno'):
        return exception.errno == 11001
             
    return isinstance(exception, paramiko.SSHException)


def try_again_on_any_exception(exception) -> bool:
    logging.error(exception)
    return True


# https://github.com/rholder/retrying


@retry(retry_on_exception=retry_if_connection_error, wait_fixed=5000, stop_max_attempt_number=30)
def safe_get(session, url, **kwargs):
    return session.get(url, **kwargs)


@retry(retry_on_exception=retry_if_connection_error, wait_fixed=5000, stop_max_attempt_number=30)
def safe_post(session: Session, url: str, **kwargs):
    return session.post(url, **kwargs)


@retry(retry_on_exception=retry_if_connection_error, wait_fixed=5000, stop_max_attempt_number=30)
def safe_patch(url, **kwargs):
    return requests.patch(url, **kwargs)

@retry(retry_on_exception=retry_if_connection_error, wait_fixed=5000, stop_max_attempt_number=30)
def safe_put(session:Session, url:str, **kwargs):
    return session.put(url, **kwargs)


class LazyDecoder(json.JSONDecoder):
    def decode(self, s, **kwargs):
        regex_replacements = [
            (re.compile(r'([^\\])\\([^\\])'), r'\1\\\\\2'),
            (re.compile(r',(\s*])'), r'\1'),
        ]
        for regex, replacement in regex_replacements:
            s = regex.sub(replacement, s)
        return super().decode(s, **kwargs)

def sendemail(host, port, username, password, subject,from_addr,to_addr, html_output, images, attachments=[]):

    msg = MIMEMultipart("related")
    msg["Subject"] = subject
    msg["To"] = username
    msg["From"] = formataddr(("no-reply", username))
    #msg["Bcc"] = to_addr don't add this, otherwise "to and cc" receivers will know who are the bcc receivers

    msg.attach(MIMEText(html_output, "html"))

    #msg.set_content(html_output, subtype='html')    

    if images:
        for image in images:
            with open(image, "rb") as fp:
                img = MIMEImage(fp.read())
            img.add_header("Content-ID", "<{}>".format(basename(image)))
            msg.attach(img)

    for f in attachments or []:
            if os.path.isfile(f):
                with open(f, "rb") as fil: 
                    ext = f.split('.')[-1:]
                    attachedfile = MIMEApplication(fil.read(), _subtype = ext)
                    attachedfile.add_header(
                        'content-disposition', 'attachment', filename=basename(f) )
                msg.attach(attachedfile)
    
    server = smtplib.SMTP_SSL(host, port)  
    #server.set_debuglevel(2)

    server.ehlo()
    #server.starttls()
    #server.ehlo()

    server.login(username, password)  
    server.sendmail(from_addr, to_addr.split(","), msg.as_string())  
    server.quit()

def sendemail_office_365(host, port, username, password, subject,from_addr,to_addr, html_output, images, attachments=[]):

    msg = MIMEMultipart("related")
    msg["Subject"] = subject
    msg["To"] = username
    msg["From"] = formataddr(("no-reply", username))
    #msg["Bcc"] = to_addr don't add this, otherwise "to and cc" receivers will know who are the bcc receivers

    msg.attach(MIMEText(html_output, "html"))

    #msg.set_content(html_output, subtype='html')    

    if images:
        for image in images:
            with open(image, "rb") as fp:
                img = MIMEImage(fp.read())
            img.add_header("Content-ID", "<{}>".format(basename(image)))
            msg.attach(img)

    for f in attachments or []:
            if os.path.isfile(f):
                with open(f, "rb") as fil: 
                    ext = f.split('.')[-1:]
                    attachedfile = MIMEApplication(fil.read(), _subtype = ext)
                    attachedfile.add_header(
                        'content-disposition', 'attachment', filename=basename(f) )
                msg.attach(attachedfile)
    
    with smtplib.SMTP(host, port) as smtp:
        smtp.starttls()
        smtp.login(username, password)
        smtp.send_message(msg,from_addr, to_addr.split(","))
        smtp.quit()

def sendemail_postmarkapp(host, port, username, password, headers, subject,from_addr,to_addr, html_output, images, attachments=[]):

    msg = MIMEMultipart("related")
    msg["Subject"] = subject
    msg["To"] = username
    msg["From"] = formataddr(("no-reply", from_addr))
    #msg["Bcc"] = to_addr don't add this, otherwise "to and cc" receivers will know who are the bcc receivers

    if headers:
        for header in headers:
            msg.add_header((header.split(":")[0]).strip(), (header.split(":")[1]).strip())

    msg.attach(MIMEText(html_output, "html"))

    #msg.set_content(html_output, subtype='html')    

    if images:
        for image in images:
            with open(image, "rb") as fp:
                img = MIMEImage(fp.read())
            img.add_header("Content-ID", "<{}>".format(basename(image)))
            msg.attach(img)

    attachments = list(set(attachments))            

    for f in attachments or []:
            if os.path.isfile(f):
                with open(f, "rb") as fil: 
                    ext = f.split('.')[-1:]
                    attachedfile = MIMEApplication(fil.read(), _subtype = ext)
                    attachedfile.add_header(
                        'content-disposition', 'attachment', filename=basename(f) )
                msg.attach(attachedfile)
    
    with smtplib.SMTP(host, port) as smtp:
        smtp.starttls()
        smtp.login(username, password)
        smtp.send_message(msg,from_addr, to_addr.split(","))
        smtp.quit()

#https://bytes.com/topic/python/answers/23091-send-email
def sendemailCDO(host, port, username, password, subject,from_addr,to_addr, html_output, images, attachments=[]):

    message = win32com.client.Dispatch("CDO.Message")
    configuration = win32com.client.Dispatch("CDO.Configuration")
    fields = configuration.Fields
    
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/SendUsing").Value = 2 # cdoSendUsingPort
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpconnectiontimeout").Value = 60
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpserver").Value = host
    
    if port:
        configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpserverport").Value = port
    
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/sendusername").Value = username 
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/sendpassword").Value = password 
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpaccountname").Value = username 
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpusessl").Value = 0 
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/sendtls").Value = 0
    configuration.Fields.Item("http://schemas.microsoft.com/cdo/configuration/smtpauthenticate").Value = 1
    
    if images:
        for image in images:
            objImage = message.AddRelatedBodyPart(os.path.abspath(image), basename(image), 0 )
            objImage.Fields("urn:schemas:mailheader:Content-ID").Value = "<{}>".format(basename(image))
            objImage.Fields.Update()
        
    configuration.Fields.Update()
    
    message.Configuration = configuration
    message.To = to_addr
    message.From = formataddr(("no-reply", from_addr)) #from_addr
    message.Subject = subject
    
    message.MimeFormatted=True
    message.HTMLBody = html_output

    message.Attachments.DeleteAll() 
    
    for f in attachments or []:
        message.AddAttachment(f)

    message.Send()

def get_last_date_of_month(year, month):
    """Return the last date of the month.
    
    Args:
        year (int): Year, i.e. 2022
        month (int): Month, i.e. 1 for January

    Returns:
        date (datetime): Last date of the current month
    """
    
    if month == 12:
        last_date = datetime(year, month, 31)
    else:
        last_date = datetime(year, month + 1, 1) + timedelta(days=-1)
    
    return last_date

def hasChildren(obj):
    try:
        return obj.Children.Count > 0
    except:
        return False
    
def findById(session, id):
    try:
        return session.findById(id)
    except:
        return None
    
def GetElementByAttribute(session, window, target, root, attribute, search):

    element = None
    cntScrollDown = getCountVerticalScrollDown(session, target)

    for i in range(0, cntScrollDown):
        
        element = innerGetElementByAttribute(findById(session, root), attribute, search)

        if (element):
            break
        else:
            pressPageDown(window)

    return element

def innerGetElementByAttribute(element, attribute, search):

    found = None

    if (getattr(element, attribute).strip() == search.strip()):
        found = element
    else:
        if hasChildren(element):
            for i in range(0, element.Children.Count - 1):
				
                found = innerGetElementByAttribute(element.Children.ElementAt(i), attribute, search)
				
                if (found):
                    break
	
    return found

def screenToXML(session, SAPRootElementId, transaction):

    SAPRootElement = session.findById(SAPRootElementId)
    XMLRootNode = xmlParser.appendChild(xmlParser.createElement(SAPRootElement.Type))
	
    enumChildrens(SAPRootElement, XMLRootNode)

    with open(transaction + ".xml", 'w') as fp:     # CONTEXT MANAGER (NO close() NEEDED)
        xmlParser.writexml(fp, addindent='\t', newl='\n', encoding="utf-8")        

def enumChildrens(SAPRootElement, XMLRootNode):
    
    i = 0

    for i in range(0, SAPRootElement.Children.Count - 1):

        SAPChildElement = SAPRootElement.Children.ElementAt(i)
        
        # Create a node
        XMLSubNode = XMLRootNode.appendChild(xmlParser.createElement(SAPChildElement.Type))

        # Attribute Name
        attrName = xmlParser.createAttribute("Name")
        attrName.Value = SAPChildElement.Name
        XMLSubNode.setAttributeNode(attrName)

        # Attribute Text
        if (len(SAPChildElement.Text) > 0):
            attrText = xmlParser.createAttribute("Text")
            attrText.Value = SAPChildElement.Text
            XMLSubNode.setAttributeNode(attrText)

        # Attribute Id
        attrId = xmlParser.createAttribute("Id")
        attrId.Value = SAPChildElement.Id
        XMLSubNode.setAttributeNode(attrId)

        # If the current object is a container, then iterate through the child elements
        #SAPChildElement.ContainerType
        #SAPRootElement.Children.Count
        if (SAPChildElement.ContainerType):
            enumChildrens(SAPChildElement, XMLSubNode)

def safe_get_dict(dictionary, *keys, default=None):
    """ Safe get items in diictionaries """
    return reduce(lambda d, key: d.get(key, default) if isinstance(d, dict) else d[key] if isinstance(key, int) else default, keys, dictionary)

def find_and_terminate_other_instance():
    current_pid = os.getpid()
    current_executable = psutil.Process(current_pid).exe()

    for process in psutil.process_iter(['pid', 'exe']):
        try:
            if process.info['pid'] != current_pid and process.info['exe'] == current_executable:
                print(f"Terminating process with PID {process.info['pid']}")
                process.terminate()
                process.wait(timeout=5)  # Wait for the process to terminate
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            print(f"Error terminating process: {e}")
            continue
    
    return False

def delete_temp_files():

    temp_dir = tempfile.gettempdir()

    logging.info(f'pasta temporária {temp_dir}')

    for nome in os.listdir(temp_dir):
        caminho = os.path.join(temp_dir, nome)
        try:
            if os.path.isfile(caminho) or os.path.islink(caminho):
                os.remove(caminho)
            elif os.path.isdir(caminho):
                shutil.rmtree(caminho)
        except Exception as e:
            logging.debug(e)
            pass        

def match(rg, texto) -> str:
    # Expressão regular fornecida
    padrao = re.compile(rg,flags=re.M|re.I)
    
    # Contar matchs
    matches = padrao.findall(texto)

    if not matches:
        return ""
    
    if len(matches) > 0:
        if type(matches[0]) is tuple and len(matches[0]) > 1:
            for item in matches[0]:
                if len(item) > 0:
                    return re.sub(r"<[^>]*>", "", re.sub(r"<[^>]*>.*?</[^>]*>", "", item))
        
    return re.sub(r"<[^>]*>", "", re.sub(r"<[^>]*>.*?</[^>]*>", "", matches[0]))

def safe_execute(func, *args, **kwargs):
    """
    Executa uma função com segurança, capturando exceções e retornando um valor padrão em caso de erro.

    :param func: A função a ser executada.
    :param args: Argumentos posicionais para a função.
    :param kwargs: Argumentos nomeados para a função.
    :return: O resultado da função, ou None se ocorrer um erro.
    """
    try:
        return func(*args, **kwargs)
    except Exception as e:
        print(f"Erro ao executar {func.__name__}: {e}")
        return None
    
def safe_xpath(dom, xpath) -> str:

    result = dom.xpath(xpath)

    return result[0] if result else ""

def safe_soup_find(soup, name, attrs, target) -> str:

    result = soup.find(name,attrs)
    
    if result and target in result.attrs:
        return result[target]
    
    if result and target:
        return getattr(result,target)
    
def bootstrap_to_logging_level(bootstrap_class):
    """
    Converte uma classe de alerta do Bootstrap em um nível do logging do Python.
    
    Args:
        bootstrap_class (str): Nome da classe do Bootstrap (ex: 'alert-warning').
    
    Returns:
        int: Nível correspondente do logging (ex: logging.WARNING).
    """
    mapping = {
        "alert-primary": logging.INFO,      # Não tem equivalente exato no logging
        "alert-secondary": logging.DEBUG,   # Geralmente usado para mensagens neutras
        "alert-success": logging.INFO,      # Sucesso pode ser INFO
        "alert-danger": logging.CRITICAL,   # 'alert-danger' é um erro crítico
        "alert-warning": logging.WARNING,   # Aviso (WARNING)
        "alert-info": logging.INFO,         # Informação (INFO)
        "alert-light": logging.DEBUG,       # Mais fraco, DEBUG
        "alert-dark": logging.DEBUG         # Pode ser DEBUG
    }
    
    return mapping.get(bootstrap_class, logging.NOTSET)  # Retorna NOTSET se não encontrar

def wait_for_load_state(page, start_wait=250):

    # page.wait_for_timeout(start_wait)

    # page.wait_for_load_state(state="domcontentloaded")

    # while page.is_visible('xpath=//div[contains(@class,"loading-message")]'):
    #     page.wait_for_timeout(start_wait/2)

    # page.wait_for_load_state(state="domcontentloaded")       

    xpath = 'xpath=//div[contains(@class,"loading-message")]'

    botao = page.locator(xpath)
    try:
        botao.wait_for(state="visible", timeout=start_wait)
    except TimeoutError:
        pass
            
    while page.is_visible(xpath):
        #page.wait_for_timeout(start_wait/2)
        try:
            botao.wait_for(state="hidden", timeout=(start_wait/2))
        except TimeoutError:
            pass                

    page.wait_for_load_state(state="domcontentloaded")    

def fill_select2(page, selector, value) -> bool:
     
    params = {"selector": selector, "value": value}
    element = page.locator(selector) 

    result = page.evaluate('''
        (params) => {
            return $(params.selector).data("select2").val();
        }
    ''', params)  

    if result == value:
        logging.debug(f'Valor {result} já estava selecionado em {params["selector"]}') 
        return True    
    
    selected_option = element.locator('option[selected]')
    if selected_option.count(): 
        if selected_option.inner_text().strip() == value:
            logging.debug(f'Valor {result} já estava selecionado em {params["selector"]}') 
            return True    
    
    page.evaluate('''
        (params) => {
            $(params.selector).data("select2").trigger("query", {"term": params.value });
        }
    ''', params)            
    
    wait_for_load_state(page)

    # # Espera ativa até o "Buscando..." desaparecer
    # while True:
        
    #     locator = page.locator(f'#select2-{element.get_attribute("id")}-results li.select2-results__option.loading-results')
    #     if not locator.count():
    #         break

    #     li_text = locator.inner_text()
    #     if not "Buscando" in li_text:
    #         break

    #     time.sleep(0.5)      

    locator = page.locator(f'//ul[@id="select2-{element.get_attribute("id")}-results"]/li[1]')
    
    if locator.count():
        page.wait_for_selector(selector, state="visible")
        locator.first.click()
        wait_for_load_state(page)

    result = page.evaluate('''
        (params) => {
            return $(params.selector).data("select2").val();
        }
    ''', params)            

    page.evaluate('''
        (params) => {
            $(params.selector).data("select2").trigger("close");
        }
    ''', params)       

    if (result):
        logging.info(f'Valor {result} selecionado em {params["selector"]}') 
    else:
        logging.error(f'Valor {value} NÃO ENCONTRADO em {params["selector"]}') 

    return (len(result or "") > 0)   

def fill_select2_by_index(page, selector, index) -> bool:
     
    params = {"selector": selector, "index": index}
    element = page.locator(selector) 
    
    page.evaluate('''
        (params) => {
            $(params.selector).data("select2").trigger("open");
        }
    ''', params)            
    
    wait_for_load_state(page)

    # # Espera ativa até o "Buscando..." desaparecer
    # buscando_selector = f'#select2-{element.get_attribute("id")}-results li.select2-results__option.loading-results'
    # while True:
    #     li_text = page.locator(buscando_selector).inner_text() if page.locator(buscando_selector).count() > 0 else ""
    #     if "Buscando" not in li_text:
    #         break
    #     time.sleep(0.5)      

    locator = page.locator(f'//ul[@id="select2-{element.get_attribute("id")}-results"]/li[{params["index"]}]')

    if locator.count():
        page.wait_for_selector(selector, state="visible")
        locator.first.click()
        wait_for_load_state(page)

    result = page.evaluate('''
        (params) => {
            return $(params.selector).data("select2").val();
        }
    ''', params)         

    page.evaluate('''
        (params) => {
            $(params.selector).data("select2").trigger("close");
        }
    ''', params)             

    if (result):
        logging.info(f'Valor {result} selecionado em {params["selector"]}') 
    else:
        logging.error(f'índice {index} NÃO ENCONTRADO em {params["selector"]}') 

    return (len(result or "") > 0)   

def get_selected_text_safe(page: Page, selector: str) -> str:
    try:
        # Primeiro tenta localizar o elemento <select>
        select_element = page.query_selector(selector)
        if not select_element:
            return ''
        
        # Dentro do select, localizar a opção selecionada
        selected_option = select_element.query_selector('option[selected]')
        if selected_option:
            return selected_option.inner_text().strip()
        
        # # Se não tiver "selected", pegar a opção atual pelo valor
        # current_value = select_element.input_value()
        # option = select_element.query_selector(f'option[value="{current_value}"]')
        # if option:
        #     return option.inner_text().strip()

        return ''
    except Exception as e:
        return ''

def sftp_exists(session, path):
    try:
        session.stat(path)
        return True
    except FileNotFoundError:
        return False  
    
def is_CPF(cpf: str) -> bool:
    cpf = ''.join(filter(str.isdigit, cpf))

    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    for i in range(9, 11):
        soma = sum(int(cpf[num]) * ((i + 1) - num) for num in range(i))
        digito = (soma * 10 % 11) % 10
        if digito != int(cpf[i]):
            return False
    return True

def is_CNPJ(cnpj: str) -> bool:
    cnpj = ''.join(filter(str.isdigit, cnpj))

    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    pesos_1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos_2 = [6] + pesos_1

    def calcula_digito(cnpj_parcial, pesos):
        soma = sum(int(d) * p for d, p in zip(cnpj_parcial, pesos))
        resto = soma % 11
        return '0' if resto < 2 else str(11 - resto)

    digito1 = calcula_digito(cnpj[:12], pesos_1)
    digito2 = calcula_digito(cnpj[:12] + digito1, pesos_2)

    return cnpj[-2:] == digito1 + digito2

def formatar_cpf_cnpj(valor) -> str:
    # Converte para string e remove qualquer caractere que não seja número
    valor_str = str(valor).strip()
    valor_str = ''.join(filter(str.isdigit, valor_str))

    if len(valor_str) <= 11:
        # Considera como CPF
        return valor_str.zfill(11)
    elif len(valor_str) <= 14:
        # Considera como CNPJ
        return valor_str.zfill(14)
    else:
        return valor_str
    
def handle_exceptions(default_return=None):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                logging.error(f"[{func.__name__}] Exception: {type(e).__name__} - {e}")
                return default_return
        return wrapper
    return decorator

def capture_aspnet_form(page: Page, form_selector="form#aspnetForm") -> tuple[dict, str]:
    """
    Captura o conteúdo do form ASP.NET e retorna:
    - form_data: dicionário com todos os campos
    - action_url: URL completa do form
    """
    result = page.evaluate(f"""
        () => {{
            const form = document.querySelector("{form_selector}");
            const data = {{}};
            const elements = form.querySelectorAll("input, select, textarea");
            for (const el of elements) {{
                if (!el.name) continue;
                if ((el.type === "checkbox" || el.type === "radio") && !el.checked) continue;
                data[el.name] = el.value;
            }}
            return {{
                data: data,
                action: form.getAttribute("action")
            }};
        }}
    """)
    form_data = result["data"]
    action = result["action"] or ""
    action_url = urljoin(page.url, action)  # resolve ./ ou paths relativos
    return form_data, action_url

def submit_postback_form(page: Page, form_data: dict, action_url: str, event_target: str = "", event_argument: str = ""):
    """
    Submete o form via Playwright context.request com EVENTTARGET/ARGUMENT
    """
    form_data["__EVENTTARGET"] = event_target
    form_data["__EVENTARGUMENT"] = event_argument

    response = page.context.request.post(
        action_url,
        data=form_data,
        headers={
            "Content-Type": "application/x-www-form-urlencoded"
        }
    )
    return response

def navigate_postback_using_form_data(page: Page, form_data: dict, action_url: str, event_target: str = "", event_argument: str = "") -> dict:
    """
    Submete o form reconstruído no DOM com os dados fornecidos.
    Realiza a navegação e retorna informações da nova página.
    """
    form_data["__EVENTTARGET"] = event_target
    form_data["__EVENTARGUMENT"] = event_argument

    fields_js = ""
    for key, value in form_data.items():
        escaped_key = key.replace('"', '\\"')
        escaped_value = value.replace('"', '\\"') if isinstance(value, str) else str(value)
        fields_js += f'''
            var input = document.createElement("input");
            input.type = "hidden";
            input.name = "{escaped_key}";
            input.value = "{escaped_value}";
            form.appendChild(input);
        '''

    try:
        with page.expect_navigation():
            page.evaluate(f"""
                () => {{
                    const form = document.createElement("form");
                    form.method = "POST";
                    form.action = "{action_url}";
                    {fields_js}
                    document.body.appendChild(form);
                    form.submit();
                }}
            """)
        
        return {
            "success": True,
            "url": page.url,
            "title": page.title(),
            "content_snippet": page.content()[:1000]
        }
    
    except TimeoutError:
        return {
            "success": False,
            "error": "Navigation timed out",
            "url": page.url,
            "title": page.title(),
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "url": page.url,
            "title": page.title(),
        }

def safely_load_workbook(*args, **kwargs):
    """
    Safely loads an Excel workbook using the same parameters as openpyxl.load_workbook,
    while handling missing, invalid, or corrupted files gracefully.

    Returns:
        openpyxl.Workbook or None: The loaded workbook, or None if an error occurs.
    """
    # Attempt to extract filename from args or kwargs
    filename = kwargs.get("filename")
    if not filename and len(args) > 0:
        filename = args[0]

    if not filename or not os.path.exists(filename):
        logging.error(f"File not found: {filename}")
        return None

    try:
        wb = load_workbook(*args, **kwargs)
        return wb
    except InvalidFileException:
        logging.error(f"Invalid or corrupted Excel file: {filename}")
    except Exception as e:
        logging.error(f"Unexpected error while loading '{filename}': {e}")

    return None

def fill_and_verify(locator: Locator, value: str, timeout: float = 3.0, max_attempts: int = 3, retry_delay: float = 0.2) -> bool:
    """
    Preenche um campo e valida se o valor foi aplicado corretamente.

    :param locator: Locator do campo de input
    :param value: Valor a preencher
    :param timeout: Tempo máximo total de espera (segundos)
    :param max_attempts: Número máximo de tentativas de preenchimento
    :param retry_delay: Tempo entre verificações (segundos)
    :return: True se o valor foi preenchido corretamente, False caso contrário
    """
    start_time = time.time()
    attempts = 0

    while time.time() - start_time < timeout and attempts < max_attempts:
        locator.fill(value)
        preenchido = locator.input_value()

        if preenchido == value:
            return True

        logging.warning(f"[Tentativa {attempts+1}] Valor incorreto: '{preenchido}', esperado: '{value}'")
        attempts += 1
        time.sleep(retry_delay)

    return False

class LMDBWrapper:
    def __init__(self, path: str = "db.lmdb", map_size: int = 4 * 1024 ** 3):
        self.path = path
        self.map_size = map_size
        self._open_env()

    def _open_env(self):
        self.env = lmdb.open(self.path, map_size=self.map_size, max_dbs=1)

    def _resize_and_retry(self, func, *args, **kwargs):
        try:
            return func(*args, **kwargs)
        except lmdb.MapFullError:
            self.map_size *= 2  # dobra o espaço
            self._open_env()    # reabre o banco com novo tamanho
            return func(*args, **kwargs)

    def insert(self, doc: dict) -> str:
        doc_id = str(uuid.uuid4())
        value = json.dumps(doc).encode()

        def op():
            with self.env.begin(write=True) as txn:
                txn.put(doc_id.encode(), value)

        self._resize_and_retry(op)
        return doc_id

    def get(self, doc_id: str) -> dict:
        with self.env.begin() as txn:
            value = txn.get(doc_id.encode())
            return json.loads(value.decode()) if value else None

    def all(self) -> List[dict]:
        with self.env.begin() as txn:
            return [json.loads(v.decode()) for _, v in txn.cursor()]

    def search(self, condition: Callable[[dict], bool]) -> List[dict]:
        with self.env.begin() as txn:
            return [
                json.loads(v.decode())
                for _, v in txn.cursor()
                if condition(json.loads(v.decode()))
            ]

    def update(self, doc_id: str, update_func: Callable[[dict], dict]) -> bool:
        def op():
            with self.env.begin(write=True) as txn:
                value = txn.get(doc_id.encode())
                if not value:
                    return False
                data = json.loads(value.decode())
                updated = update_func(data)
                txn.put(doc_id.encode(), json.dumps(updated).encode())
                return True

        return self._resize_and_retry(op)

    def remove(self, doc_id: str) -> bool:
        def op():
            with self.env.begin(write=True) as txn:
                return txn.delete(doc_id.encode())

        return self._resize_and_retry(op)

    def usage_stats(self) -> dict:
        stat = self.env.stat()
        info = self.env.info()
        used = stat["psize"] * (info["last_pgno"] + 1)
        return {
            "used_bytes": used,
            "used_MB": round(used / 1024**2, 2),
            "allocated_MB": round(self.map_size / 1024**2, 2),
        }
    
    def count_where(self, condicao: Callable[[dict], bool]) -> int:
        """
        Conta quantos documentos satisfazem a condição fornecida.
        :param condicao: Função que recebe um dict e retorna True/False
        :return: Total correspondente
        """
        count = 0
        with self.env.begin() as txn:
            cursor = txn.cursor()
            for _, v in cursor:
                try:
                    doc = json.loads(v.decode())
                    if condicao(doc):
                        count += 1
                except Exception as e:
                    print(f"[count_where] Erro ao processar registro: {e}")
        return count

    def delete_where(self, condicao: Callable[[dict], bool], batch_size: int = 1000) -> int:
        """
        Remove documentos que satisfazem a condição, em batch.
        :param condicao: Função que recebe um dict e retorna True/False
        :param batch_size: Número de exclusões por transação (segurança e performance)
        :return: Total de registros removidos
        """
        deleted = 0
        to_delete = []

        # Coleta chaves a serem removidas
        with self.env.begin() as txn:
            cursor = txn.cursor()
            for k, v in cursor:
                try:
                    doc = json.loads(v.decode())
                    if condicao(doc):
                        to_delete.append(k)
                except Exception as e:
                    print(f"[delete_where] Erro ao processar registro: {e}")

        # Remove em batches
        for i in range(0, len(to_delete), batch_size):
            batch = to_delete[i:i + batch_size]
            with self.env.begin(write=True) as txn:
                for k in batch:
                    txn.delete(k)
                    deleted += 1

        return deleted


def transform_timestamp_to_datetime(timestamp: int) -> datetime:
    """
    Converte um timestamp Unix (segundos desde 1970-01-01) para um objeto datetime.

    Args:
        timestamp (int): Timestamp Unix.

    Returns:
        datetime: Objeto datetime correspondente ao timestamp.
    """
    format_timestamp = int(timestamp)
    date = datetime.fromtimestamp(format_timestamp / 1000)
    return date.strftime("%d/%m/%Y")

def safe_get_phone_number(ocurrence_data: dict, form_data: dict) -> str:
    beneficiario = ocurrence_data.get("beneficiario")
    telefone_contato = form_data.get("telefone-contato")

    if isinstance(beneficiario, dict) and "telefone" in beneficiario:
        return beneficiario["telefone"]
    elif isinstance(telefone_contato, str) and telefone_contato.strip():
        return telefone_contato
    else:
        return ""


def safe_get_label_in_array(data: dict, key: str) -> str:
    """
    Pega o primeiro label da lista dentro da chave `key` no dicionário `data`.
    Se não existir ou estiver vazio, retorna uma mensagem padrão.
    """

    items = data.get(key) or [{}]
    if isinstance(items, list) and len(items) > 0:
        label = items[0].get("label")
        if isinstance(label, list) and len(label) > 0:
            return label[0]
    return ""


def safe_get_text(data: dict, key: str) -> str:
    if not isinstance(data, dict):
        return ""
    value = data.get(key)
    return value if value else ""


def safe_supplier_document(data: dict) -> str:
    """
    Obtém o documento do fornecedor de forma segura.
    """
    document_type = safe_get_label_in_array(data, "tipo-prestador")

    if document_type == "Jurídica":
        return safe_get_text(data, "cpf-cnpj-do-prestador")

    return safe_get_text(data, "cpf")


def get_url_payment_data(service_tab: list):
    id_form = ""
    id_form_data = ""
    url = ""

    for service in service_tab:
        if service.get("nome").lower() == "pagamento":
            form_data = service.get("formDatas", {})

            if not form_data:
                return None

            id_form_data = form_data[0]
            id_form = service.get("id_form")
            url = f"https://api.mosiaomnichannel.com.br/privado/omni/form_builder/v1/gan/saudeams/form/{id_form}/data/{id_form_data}"

            return url


def get_payment_data(url: str, session: Session):
    response = session.get(url)

    if not response.ok:
        logging.error(
            f"Erro ao obter dados de pagamento: {response.status_code} - {response.text}")
        return None

    payment_info = response.json()
    return payment_info.get("data")


def get_situation_options(url, session:Session):
    status_list = []

    response = session.get(url)

    if not response.ok:
        logging.error(
            f"Erro ao obter opções de situação: {response.status_code} - {response.text}")
        return None

    situation_info = response.json()
    for item in situation_info.get("data", []):
        status_list.append(item.get("id_status"))
        status_list.append(item.get("descricao_interna"))

    return status_list

@timeit
@handle_exceptions(default_return=False)
@retry(retry_on_exception=try_again_on_any_exception, wait_fixed=10000, stop_max_attempt_number=5)
def login_mobile_saude(session:Session) -> bool:

    payload = {
        "domain": config["mobilesaude"]["domain"],
        "password": config["mobilesaude"]["password"],
        "user": config["mobilesaude"]["username"]
    }

    response = safe_post(
        session,
        global_parameters["mobilesaude.submit_login"], data=payload, verify=False)

    if not response.ok:
        logging.error(f"Erro no login: {response.status_code}")
        return False

    token = safe_get_text(response.json(), "token")

    if not token:
        logging.error(f"Token não encontrado na resposta do login")
        return False
    
    env = global_parameters["env"]
    env_filter = global_parameters[f"mobilesaude.env_filter_{env}"]
    
    session.headers.update({
        "instancia_aplicacao": str(env_filter),
        "Authorization": f"{token}",
    })

    post_status = safe_post(
        session,
        global_parameters["mobilesaude.change_attendant_status"], 
        data={"status": global_parameters["mobilesaude.attendant_status"]}, 
        verify=False
    )

    if not post_status.ok:
        logging.error(f"Erro ao atualizar status: {post_status.status_code}")
        return False

    logging.info(f'Login com sucesso {config["mobilesaude"]["username"]}...')
    logging.info(f'Navegando para SAUDE AMS...')

    return True


def convert_date(date: str) -> str:
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        return date_obj.strftime("%d/%m/%Y")
    except ValueError:
        return date
    
@timeit
def reload_old_files():

    def find_xlsx_files(base_directory):
        xlsx_files = []

        for root, dirs, files in os.walk(base_directory):
            for file in files:
                if file.lower().startswith('protocolos') and file.lower().endswith('.xlsx'):
                    full_path = os.path.join(root, file)
                    xlsx_files.append(full_path)

        return xlsx_files

    logging.info(f'Recarregando arquivos de processamentos anteriores...')

    worksheets = find_xlsx_files(f'{path}\\log')
    worksheets_count = 0
    worksheets_total = len(worksheets)

    column_names = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="labels", header=None).iloc[:, 0].tolist()
    column_ids = pd.read_excel("./data/parameters.xlsx", engine='openpyxl', sheet_name="ids", header=None).iloc[:, 0].tolist()

    db = LMDBWrapper()

    logging.info(db.usage_stats())

    for worksheet in worksheets:

        worksheets_count += 1         

        logging.info(f'Reprocessando planilha {os.path.basename(worksheet)} {worksheets_count}/{worksheets_total}...')

        result = db.count_where(lambda d: d["file_id"] == os.path.basename(worksheet))     
        
        #
        wb = safely_load_workbook(worksheet, read_only=True)
        if not wb:
            continue
        
        ws = wb.active
        protocols_total = ws.max_row -1

        if result == protocols_total:
            continue

        colunas = {cell.value: cell.column for cell in ws[1]}
        new_colunas = {}

        for chave, valor in colunas.items():
            if chave in column_names:
                index = column_names.index(chave)
                nova_chave = column_ids[index]
            else:
                nova_chave = chave  # mantém original se não encontrar
            new_colunas[nova_chave] = valor - 1 

        colunas = new_colunas
        protocols_count = 0

        db.delete_where(lambda doc: doc["file_id"] == os.path.basename(worksheet))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

            protocols_count += 1         

            logging.info(f'Incluindo protocolo {row[colunas.get("protocol_id")].value} no BD {protocols_count}/{protocols_total}...')

            record = {
                'file_id': os.path.basename(worksheet),
                'env': row[colunas.get("env")].value if colunas.get("env") else ("prd" if int(row[colunas.get("refund_id")].value) > 999999 else "hml"),
                'refund_id': row[colunas.get("refund_id")].value,
                'protocol_id': row[colunas.get("protocol_id")].value,
                'protocol_date': row[colunas.get("protocol_date")].value,
                'status_id': row[colunas.get("status_id")].value,
                'status_desc': row[colunas.get("status_desc")].value,
                'refund_type': row[colunas.get("refund_type")].value,
                'refund_qty': row[colunas.get("refund_qty")].value,
                'refund_value': row[colunas.get("refund_value")].value,
                'card': row[colunas.get("card")].value,
                'user': row[colunas.get("user")].value if colunas.get("user") else None,
                'holder_name': row[colunas.get("holder_name")].value,
                'holder_cpf': row[colunas.get("holder_cpf")].value,
                'phone_number': row[colunas.get("phone_number")].value,
                'plan': row[colunas.get("plan")].value,
                'payment_day': row[colunas.get("payment_day")].value,
                'payment_type': row[colunas.get("payment_type")].value,
                'lot': row[colunas.get("lot")].value,
                'expense_id': row[colunas.get("expense_id")].value,
                'expense_status': row[colunas.get("expense_status")].value,
                'supplier_id': row[colunas.get("supplier_id")].value,
                'supplier_name': row[colunas.get("supplier_name")].value,
                'supplier_state': row[colunas.get("supplier_state")].value,
                'supplier_city': row[colunas.get("supplier_city")].value,
                'expense_date': row[colunas.get("expense_date")].value,
                'expense_nf': row[colunas.get("expense_nf")].value,
                'guide_number': row[colunas.get("guide_number")].value if colunas.get("guide_number") else None,
                'PEG': row[colunas.get("PEG")].value,
                'notes': row[colunas.get("notes")].value,
                'assigned': row[colunas.get("assigned")].value,
                'comment': row[colunas.get("comment")].value,
                'complement': row[colunas.get("complement")].value if colunas.get("complement") else None,
            }

            db.insert(record)

    logging.info(f'Recarregando arquivos de processamentos anteriores... OK')

    return None


def get_occurrence_data(session: Session, occurrence: str) -> dict:
    url = global_parameters["mobilesaude.occurrence_details"] % occurrence
    response = session.get(url)

    if not response.ok:
        return None

    return response.json().get("data")


def get_occurrence_by_protocol(session: Session, protocol_id: str) -> dict:
    response = session.get(global_parameters["mobilesaude.query_protocolos"], headers={
        "search": protocol_id
    })

    if not response.ok or not response.json().get("status"):
        return None

    return response.json().get("data")[0]


def get_attendance_data(session: Session) -> dict :
    url = global_parameters["mobilesaude.get_attendants"]
    occurrence_type = global_parameters["mobilesaude_occurrence_type_filter"]
    headers = {"idtipoocorrencia": str(occurrence_type)}
    response = session.get(url, headers=headers)

    if not response.ok or not response.json().get("status"):
        return None

    return response.json().get("data")